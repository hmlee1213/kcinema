# app.py — 서울 독립극장 시간표 서비스 (PostgreSQL)
import io, os, threading, logging, time, re
from datetime import datetime, date, timedelta
from flask import Flask, jsonify, request, send_file, session, redirect, url_for, render_template_string
import requests
from bs4 import BeautifulSoup
import psycopg2
from psycopg2.extras import RealDictCursor

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

app = Flask(__name__, template_folder="templates")

# SECRET_KEY: Railway 고유값 조합 자동생성 (별도 설정 불필요)
# 직접 설정한 값이 있으면 그걸 우선 사용
app.secret_key = os.environ.get("SECRET_KEY") or (
    os.environ.get("RAILWAY_PROJECT_ID", "local") + "-" +
    os.environ.get("RAILWAY_SERVICE_ID", "dev")
)

# ADMIN_PASSWORD: Railway Variables에서 반드시 직접 설정
# 설정 안 하면 어드민 페이지 자체가 비활성화됨
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "")

DATABASE_URL = os.environ.get("DATABASE_URL")

# ── 브루트포스 방어 (IP별 시도 횟수, 메모리 저장) ───────
# { ip: {"count": int, "locked_until": datetime | None} }
_login_attempts: dict = {}
MAX_ATTEMPTS    = 10
LOCKOUT_MINUTES = 30

def _get_ip():
    return request.headers.get("X-Forwarded-For", request.remote_addr or "").split(",")[0].strip()

def _is_locked(ip):
    rec = _login_attempts.get(ip)
    if not rec: return False
    if rec["locked_until"] and datetime.now() < rec["locked_until"]:
        return True
    if rec["locked_until"] and datetime.now() >= rec["locked_until"]:
        # 잠금 해제 — 카운트 초기화
        _login_attempts.pop(ip, None)
    return False

def _record_failure(ip):
    rec = _login_attempts.setdefault(ip, {"count": 0, "locked_until": None})
    rec["count"] += 1
    if rec["count"] >= MAX_ATTEMPTS:
        rec["locked_until"] = datetime.now() + timedelta(minutes=LOCKOUT_MINUTES)
        log.warning(f"관리자 로그인 잠금: {ip} ({MAX_ATTEMPTS}회 실패)")

def _clear_attempts(ip):
    _login_attempts.pop(ip, None)

CINEMA_ORDER = [
    "KT&G상상마당시네마","KU시네마테크","더숲아트시네마",
    "라이카시네마","서울아트시네마","씨네큐브",
    "아리랑시네센터","아트나인","아트하우스모모","에무시네마",
    "서울영화센터","한국영상자료원"
]
FREE_CINEMAS = ["서울영화센터","한국영상자료원"]

CINEMAS = [
    {"name":"KU시네마테크","source":"moviee","t_id":"121"},
    {"name":"KT&G상상마당시네마","source":"moviee","t_id":"123"},
    {"name":"서울아트시네마","source":"seoulart"},
    {"name":"한국영상자료원","source":"kofa"},
    {"name":"라이카시네마","source":"dtryx","brand":"spacedog","cinema_cd":"000072"},
    {"name":"씨네큐브","source":"dtryx","brand":"cinecube","cinema_cd":"000003"},
    {"name":"더숲아트시네마","source":"dtryx","brand":"indieart","cinema_cd":"000065"},
    {"name":"아트하우스모모","source":"dtryx","brand":"indieart","cinema_cd":"000067"},
    {"name":"서울영화센터","source":"dtryx","brand":"seoulcc","cinema_cd":"000160"},
    {"name":"아리랑시네센터","source":"dtryx","brand":"etc","cinema_cd":"000088"},
    {"name":"에무시네마","source":"dtryx","brand":"indieart","cinema_cd":"000069"},
    {"name":"아트나인","source":"dtryx","brand":"etc","cinema_cd":"000162"},
]

DTRYX_CGID = "FE8EF4D2-F22D-4802-A39A-D58F23A29C1E"

# ── DB ───────────────────────────────────────────────
def get_db():
    conn = psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)
    return conn

def ensure_db():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS screenings (
            cinema TEXT, movie TEXT, start_dt TEXT, end_dt TEXT,
            runtime INTEGER, screen TEXT, source TEXT,
            show_type TEXT, program TEXT, movie_url TEXT,
            PRIMARY KEY(cinema, start_dt, screen)
        )
    """)
    cur.execute("ALTER TABLE screenings ADD COLUMN IF NOT EXISTS movie_url TEXT")
    cur.execute("ALTER TABLE screenings ADD COLUMN IF NOT EXISTS movie_cd TEXT")
    cur.execute("""
        CREATE TABLE IF NOT EXISTS meta (key TEXT PRIMARY KEY, value TEXT)
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS cinemas (
            name TEXT PRIMARY KEY,
            address TEXT DEFAULT '',
            url TEXT DEFAULT '',
            phone TEXT DEFAULT '',
            description TEXT DEFAULT '',
            is_free BOOLEAN DEFAULT FALSE,
            note TEXT DEFAULT '',
            updated_at TIMESTAMP DEFAULT NOW()
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS movies (
            id SERIAL PRIMARY KEY,
            title TEXT UNIQUE NOT NULL,
            title_en TEXT DEFAULT '',
            director TEXT DEFAULT '',
            country TEXT DEFAULT '',
            year INTEGER,
            runtime INTEGER,
            synopsis TEXT DEFAULT '',
            poster_url TEXT DEFAULT '',
            kobis_url TEXT DEFAULT '',
            kofa_url TEXT DEFAULT '',
            note TEXT DEFAULT '',
            updated_at TIMESTAMP DEFAULT NOW()
        )
    """)
    # 극장 기본 정보 초기 삽입 (없는 경우만)
    cinema_data = [
        ("KU시네마테크","서울 광진구 능동로 120 건국대학교 예술문화관 지하","https://www.kucine.com","02-450-3551","건국대학교 내 시네마테크. 독립·예술영화 및 특별전 중심.",False,""),
        ("KT&G상상마당시네마","서울 마포구 어울마당로 65","https://www.sangsangmadang.com","02-330-6220","홍대 KT&G상상마당 내 위치. 독립·예술·기획 상영 전문.",False,""),
        ("서울아트시네마","서울 종로구 삼일대로 428 낙원상가 4층","https://www.cinematheque.seoul.kr","02-741-9782","한국 최초 시네마테크 전문관. 고전·회고전·특별전 중심.",False,""),
        ("한국영상자료원","서울 마포구 월드컵북로 400","https://www.koreafilm.or.kr","02-3153-2001","한국 영화문화유산 보존기관. 무료 상영 및 아카이브 특별전.",True,"무료 상영"),
        ("라이카시네마","서울 마포구 와우산로29나길 20","https://www.likacinema.com","","홍대 인근 소규모 독립영화 전문관.",False,""),
        ("씨네큐브","서울 종로구 광화문 KT빌딩 지하 2층","https://www.cinecube.co.kr","02-2002-7782","광화문 위치. 예술·독립영화 전문 상영관.",False,""),
        ("더숲아트시네마","서울 강남구 언주로 154","https://www.thesoop.com","","강남 예술영화 전문관. 상영 시작 10분 지연.",False,"정시 기준 10분 지연 상영"),
        ("아트하우스모모","서울 서대문구 신촌역로 22 메가박스 신촌점 5층","https://www.arthousemomo.com","02-332-5558","신촌 메가박스 내 독립·예술영화 전문관.",False,""),
        ("서울영화센터","서울 중구 퇴계로 213","https://www.sfc.kr","02-2285-0562","서울시 지원 독립·예술영화 전용관. 무료 상영.",True,"무료 상영"),
        ("아리랑시네센터","서울 성북구 아리랑로 82","https://www.arirancine.co.kr","02-936-5008","성북구 위치 지역 예술·독립영화관.",False,""),
        ("에무시네마","서울 종로구 창경궁로 89 EMMU 지하","https://www.emucinema.com","","종로 위치 소규모 독립영화 전문관.",False,""),
        ("아트나인","서울 강남구 선릉로 221 강남 S타워 지하","https://www.artnine.co.kr","02-563-5670","강남 예술영화 전문 상영관.",False,""),
    ]
    for c in cinema_data:
        cur.execute(
            "INSERT INTO cinemas (name,address,url,phone,description,is_free,note) VALUES (%s,%s,%s,%s,%s,%s,%s) ON CONFLICT (name) DO NOTHING",
            c
        )
    cur.execute("""
        CREATE TABLE IF NOT EXISTS recommended (
            id SERIAL PRIMARY KEY,
            title TEXT UNIQUE NOT NULL,
            is_rec BOOLEAN DEFAULT TRUE,
            comment TEXT DEFAULT '',
            awards TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT NOW()
        )
    """)
    cur.execute("ALTER TABLE recommended ADD COLUMN IF NOT EXISTS is_rec BOOLEAN DEFAULT TRUE")
    cur.execute("ALTER TABLE recommended ADD COLUMN IF NOT EXISTS comment TEXT DEFAULT ''")
    cur.execute("ALTER TABLE recommended ADD COLUMN IF NOT EXISTS awards TEXT DEFAULT ''")
    # awards에 movie_info 테이블 통합 - 수상내역은 recommended에서 관리
    cur.execute("""
        CREATE TABLE IF NOT EXISTS events (
            id SERIAL PRIMARY KEY,
            title TEXT NOT NULL,
            cinema TEXT NOT NULL,
            event_date DATE NOT NULL,
            start_time TEXT DEFAULT '',
            end_time TEXT DEFAULT '',
            description TEXT DEFAULT '',
            is_free BOOLEAN DEFAULT FALSE,
            url TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT NOW()
        )
    """)
    cur.execute("ALTER TABLE cinemas ADD COLUMN IF NOT EXISTS price_info TEXT DEFAULT ''")
    cur.execute("ALTER TABLE cinemas ADD COLUMN IF NOT EXISTS price_table JSONB DEFAULT '{}'")
    cur.execute("""
        CREATE TABLE IF NOT EXISTS posters (
            title TEXT PRIMARY KEY,
            data  BYTEA NOT NULL,
            mime  TEXT DEFAULT 'image/jpeg',
            updated_at TIMESTAMP DEFAULT NOW()
        )
    """)
    conn.commit(); cur.close(); conn.close()

def save_to_db(rows):
    if not rows: return
    conn = get_db(); cur = conn.cursor()
    for r in rows:
        cur.execute("""
            INSERT INTO screenings (cinema,movie,start_dt,end_dt,runtime,screen,source,show_type,program)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT (cinema, start_dt, screen) DO UPDATE SET
                movie=EXCLUDED.movie, end_dt=EXCLUDED.end_dt,
                runtime=EXCLUDED.runtime, source=EXCLUDED.source,
                show_type=EXCLUDED.show_type, program=EXCLUDED.program
        """, (
            r["cinema"], r["movie"],
            str(r["start_dt"]) if r["start_dt"] else None,
            str(r["end_dt"])   if r["end_dt"]   else None,
            r["runtime"], r["screen"], r["source"],
            r.get("show_type",""), r.get("program","")
        ))
    cur.execute("""
        INSERT INTO meta (key, value) VALUES ('last_updated', %s)
        ON CONFLICT (key) DO UPDATE SET value=EXCLUDED.value
    """, (datetime.now().isoformat(),))
    conn.commit(); cur.close(); conn.close()

# ── 공통 함수 (getdb.py 원본) ─────────────────────────
def make_datetime(date_obj, time_str):
    if not time_str: return None
    try:
        return datetime.combine(date_obj, datetime.strptime(time_str,"%H:%M").time())
    except:
        return None

def compute_end_dt(start_dt, runtime):
    return start_dt+timedelta(minutes=runtime) if start_dt and runtime else None

def compute_runtime(start_time, end_time):
    try:
        s = datetime.strptime(start_time, "%H:%M")
        e = datetime.strptime(end_time, "%H:%M")
        return int((e-s).total_seconds()//60)
    except:
        return None

# ── fetch 함수들 (getdb.py 원본 그대로) ──────────────
def fetch_dtryx(cinema, start_date, days=14):
    rows=[]
    for i in range(days):
        day=start_date+timedelta(days=i)
        day_str=day.strftime("%Y-%m-%d")
        params={"cgid":DTRYX_CGID,"BrandCd":cinema["brand"],"CinemaCd":cinema["cinema_cd"],
                "PlaySDT":day_str,"_":int(time.time()*1000)}
        headers={"User-Agent":"Mozilla/5.0","X-Requested-With":"XMLHttpRequest"}
        try:
            data=requests.get("https://www.dtryx.com/cinema/showseq_list.do", params=params, headers=headers, timeout=10).json()
        except:
            continue
        for item in data.get("Showseqlist",[]):
            start=item.get("StartTime"); end=item.get("EndTime")
            runtime=compute_runtime(start,end)
            start_dt=make_datetime(day,start)
            end_dt=make_datetime(day,end) if end else compute_end_dt(start_dt,runtime)
            special_fields=[item.get(f,"").strip() for f in ["ScreenTypeNmNat","PlayTimeTypeNm","DisplayTypeDetailNm","ScreeningInfoNat"]]
            show_type="일반" if all(f in ["","일반"] for f in special_fields) else " / ".join(f for f in special_fields if f not in ["","일반"])
            screen_name=item.get("ScreenNmNat") or item.get("ScreenNm") or ""
            program=item.get("ProgramName","").strip() if item.get("ProgramName") else ""
            rows.append({"cinema":cinema["name"],"movie":item.get("MovieNmNat"),
                         "start_dt":start_dt,"end_dt":end_dt,"runtime":runtime,
                         "screen":screen_name,"source":"dtryx",
                         "show_type":show_type,"program":program})
    return rows

def fetch_moviee(cinema,start_date,days=14):
    rows=[]
    url="https://moviee.co.kr/api/TicketApi/GetPlayTimeList"
    for i in range(days):
        day=start_date+timedelta(days=i)
        params={"tId":cinema["t_id"],"playDt":day.strftime("%Y-%m-%d")}
        try:
            data=requests.get(url, params=params, timeout=10).json()
        except:
            continue
        if data.get("ResCd")!="00": continue
        for item in data["ResData"]["Table"]:
            start=item["PLAY_TIME"]; end=item["END_TIME"]
            start=f"{start[:2]}:{start[2:]}" if start else None
            end=f"{end[:2]}:{end[2:]}" if end else None
            runtime=compute_runtime(start,end)
            start_dt=make_datetime(day,start)
            end_dt=make_datetime(day,end) if end else compute_end_dt(start_dt,runtime)
            movie_name=item["M_NM"]
            show_type=""
            if "(" in movie_name and ")" in movie_name:
                main_title=movie_name.split("(")[0].strip()
                show_type=movie_name.split("(")[1].replace(")","").strip()
            else:
                main_title=movie_name.strip()
            screen_name=item.get("ROOM_NM","")
            rows.append({"cinema":cinema["name"],"movie":main_title,
                         "start_dt":start_dt,"end_dt":end_dt,"runtime":runtime,
                         "screen":screen_name,"source":"moviee",
                         "show_type":show_type,"program":""})
    return rows

def fetch_seoulart(cinema):
    rows=[]
    url="https://www.cinematheque.seoul.kr/bbs/content.php?co_id=timetable"
    try:
        res=requests.get(url, timeout=10)
        soup=BeautifulSoup(res.text,"html.parser")
    except:
        return rows
    year=datetime.today().year
    theater_name=cinema["name"]
    tables=soup.select("table")
    for table in tables:
        date_rows=table.select("tr.date-label")
        for dr in date_rows:
            tds=dr.find_all("td")
            col_dates=[]
            for td in tds:
                txt=td.get_text(strip=True)
                if txt:
                    m,d,_=txt.split(".")
                    col_dates.append(datetime(year,int(m),int(d)))
                else:
                    col_dates.append(None)
            next_tr=dr.find_next_sibling()
            while next_tr and "event" in next_tr.get("class",[]):
                for idx, td in enumerate(next_tr.find_all("td")):
                    date_obj=col_dates[idx] if idx<len(col_dates) else None
                    if not date_obj: continue
                    for link in td.find_all("a"):
                        try:
                            start=link.find("strong").text.strip()
                            title_tag=link.find_all("p")[1]
                            title=re.sub(r"\(\d+min\)","",title_tag.text).strip()
                            runtime_match=re.search(r"\((\d+)min\)",title_tag.text)
                            runtime=int(runtime_match.group(1)) if runtime_match else None
                            start_dt=make_datetime(date_obj,start)
                            end_dt=compute_end_dt(start_dt,runtime)
                            rows.append({"cinema":theater_name,"movie":title,
                                         "start_dt":start_dt,"end_dt":end_dt,"runtime":runtime,
                                         "screen":"","source":"seoulart",
                                         "show_type":"","program":""})
                        except:
                            continue
                next_tr=next_tr.find_next_sibling()
    return rows

def fetch_kofa(cinema):
    rows=[]
    url="https://www.koreafilm.or.kr/cinematheque/schedule"
    try:
        res=requests.get(url, timeout=10)
        soup=BeautifulSoup(res.text,"html.parser")
    except:
        return rows
    today=datetime.today()
    month=today.month
    prev_day=None
    year=today.year
    blocks=soup.find_all("dl","list-day-1")
    for block in blocks:
        date_tag=block.find("dt","txt-day")
        if not date_tag: continue
        txt=date_tag.text.strip()
        if "." not in txt: continue
        day=int(txt.split(".")[0])
        if prev_day and day<prev_day: month+=1
        prev_day=day
        date_obj=datetime(year,month,day)
        for s in block.select("ul.list-detail-1"):
            start_tag=s.select_one(".txt-time")
            start=start_tag.get_text(strip=True) if start_tag else None
            runtime_tag=s.select_one(".min")
            if runtime_tag:
                for strong in runtime_tag.find_all("strong"):
                    strong.decompose()
                runtime_text=runtime_tag.get_text(strip=True).replace("분","")
                runtime=int(runtime_text) if runtime_text.isdigit() else None
            else:
                runtime=None
            end_dt=compute_end_dt(make_datetime(date_obj,start),runtime)
            title_tag=s.select_one(".txt-1 a")
            title=title_tag.get_text(strip=True) if title_tag else ""
            screen_tag=s.select_one(".txt-room")
            screen=screen_tag.get_text(strip=True) if screen_tag else ""
            type_tag=s.select_one(".fomat")
            show_type=type_tag.get_text(strip=True)[4:] if type_tag else ""
            program_tag=s.select_one(".layer-txt-1")
            program=program_tag.get_text(strip=True) if program_tag else ""
            rows.append({
                "cinema":cinema["name"], "movie":title,
                "start_dt":make_datetime(date_obj,start), "end_dt":end_dt,
                "runtime":runtime, "screen":screen, "source":"kofa",
                "show_type":show_type, "program":program
            })
    return rows

# ── 크롤링 ───────────────────────────────────────────
def run_crawl():
    log.info("크롤링 시작")
    start_time = time.time()
    all_rows = []
    today = datetime.today()
    for cinema in CINEMAS:
        src = cinema["source"]
        try:
            if src=="dtryx":      rows=fetch_dtryx(cinema,today,days=14)
            elif src=="moviee":   rows=fetch_moviee(cinema,today,days=14)
            elif src=="seoulart": rows=fetch_seoulart(cinema)
            elif src=="kofa":     rows=fetch_kofa(cinema)
            else: rows=[]
            all_rows += rows
            log.info(f"  {cinema['name']} 수집 완료: {len(rows)}건")
        except Exception as e:
            log.error(f"  {cinema['name']} 오류: {e}")
    all_rows = sorted(all_rows, key=lambda x: (x['start_dt'] or datetime.max, x['end_dt'] or datetime.max))
    save_to_db(all_rows)
    log.info(f"총 {len(all_rows)}건 저장 완료, 처리시간: {time.time()-start_time:.2f}초")
    return len(all_rows)



# ── API ──────────────────────────────────────────────
@app.route("/")
def index():
    with open(os.path.join(app.template_folder, "index.html"), encoding="utf-8") as f:
        return f.read()

@app.route("/api/screenings")
def api_screenings():
    df = request.args.get("date_from", date.today().isoformat())
    dt = request.args.get("date_to",   date.today().isoformat())
    tf = request.args.get("time_from", "00:00")
    tt = request.args.get("time_to",   "23:59")
    cinemas = request.args.getlist("cinema")
    mq = request.args.get("movie", "").strip()
    conn = get_db(); cur = conn.cursor()
    sql = """
        SELECT s.*, COALESCE(m.poster_url, '') AS poster_url
        FROM screenings s
        LEFT JOIN movies m ON m.title = s.movie
        WHERE s.start_dt::date BETWEEN %s AND %s
          AND s.start_dt::time >= %s::time
          AND s.start_dt::time <= %s::time
    """
    params = [df, dt, tf, tt]
    if cinemas:
        sql += f" AND s.cinema = ANY(%s)"; params.append(cinemas)
    if mq:
        sql += " AND s.movie ILIKE %s"; params.append(f"%{mq}%")
    sql += " ORDER BY s.start_dt ASC"
    cur.execute(sql, params)
    rows = [dict(r) for r in cur.fetchall()]
    cur.close(); conn.close()
    return jsonify(rows)

@app.route("/api/movies")
def api_movies():
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT title, director, synopsis, poster_url FROM movies WHERE poster_url != '' ORDER BY title")
    rows = [dict(r) for r in cur.fetchall()]
    cur.close(); conn.close()
    return jsonify(rows)

@app.route("/api/recommended", methods=["GET"])
def api_recommended_get():
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT id, title, is_rec, comment, awards, created_at FROM recommended WHERE is_rec=TRUE ORDER BY created_at DESC")
    rows = [dict(r) for r in cur.fetchall()]
    cur.close(); conn.close()
    return jsonify(rows)

@app.route("/api/recommended", methods=["POST"])
def api_recommended_post():
    if not session.get("admin"):
        return jsonify({"error": "unauthorized"}), 401
    data = request.get_json()
    title   = (data.get("title") or "").strip()
    is_rec  = bool(data.get("is_rec", True))
    comment = (data.get("comment") or "").strip()
    awards  = (data.get("awards") or "").strip()
    if not title:
        return jsonify({"error": "title required"}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute("""
        INSERT INTO recommended (title, is_rec, comment, awards)
        VALUES (%s, %s, %s, %s)
        ON CONFLICT (title) DO UPDATE SET
            is_rec=EXCLUDED.is_rec, comment=EXCLUDED.comment, awards=EXCLUDED.awards
    """, (title, is_rec, comment, awards))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True})

@app.route("/api/recommended/<int:rid>", methods=["DELETE"])
def api_recommended_delete(rid):
    if not session.get("admin"):
        return jsonify({"error": "unauthorized"}), 401
    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM recommended WHERE id=%s", (rid,))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True})

@app.route("/api/cinemas")
def api_cinemas():
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT DISTINCT cinema FROM screenings")
    names = [r["cinema"] for r in cur.fetchall()]
    cur.close(); conn.close()
    ordered = [c for c in CINEMA_ORDER if c in names] + [c for c in names if c not in CINEMA_ORDER]
    return jsonify(ordered)

@app.route("/api/stats")
def api_stats():
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT COUNT(*) AS c FROM screenings"); total = cur.fetchone()["c"]
    cur.execute("SELECT COUNT(DISTINCT cinema) AS c FROM screenings"); cc = cur.fetchone()["c"]
    cur.execute("SELECT COUNT(DISTINCT movie) AS c FROM screenings"); mc = cur.fetchone()["c"]
    cur.execute("SELECT value FROM meta WHERE key='last_updated'"); lu = cur.fetchone()
    cur.execute("SELECT MIN(start_dt::date) AS mn, MAX(start_dt::date) AS mx FROM screenings WHERE start_dt IS NOT NULL")
    dr = cur.fetchone()
    cur.close(); conn.close()
    return jsonify({"total": total, "cinemas": cc, "movies": mc,
                    "last_updated": lu["value"] if lu else None,
                    "date_min": str(dr["mn"]) if dr and dr["mn"] else None,
                    "date_max": str(dr["mx"]) if dr and dr["mx"] else None})

@app.route("/api/cinema/<name>")
def api_cinema_detail(name):
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT * FROM cinemas WHERE name = %s", (name,))
    row = cur.fetchone()
    cinema = dict(row) if row else {"name": name}
    cur.execute(
        "SELECT * FROM screenings WHERE cinema = %s AND start_dt::date >= CURRENT_DATE ORDER BY start_dt ASC",
        (name,)
    )
    screenings = [dict(r) for r in cur.fetchall()]
    cur.execute(
        "SELECT MIN(start_dt::date) AS mn, MAX(start_dt::date) AS mx FROM screenings WHERE cinema = %s AND start_dt IS NOT NULL",
        (name,)
    )
    dr = cur.fetchone()
    cur.close(); conn.close()
    return jsonify({
        "cinema": cinema,
        "screenings": screenings,
        "date_max": str(dr["mx"]) if dr and dr["mx"] else None,
    })

@app.route("/api/cinemas/all")
def api_cinemas_all():
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT * FROM cinemas ORDER BY name")
    rows = [dict(r) for r in cur.fetchall()]
    cur.close(); conn.close()
    return jsonify(rows)

@app.route("/cinema/<name>")
def cinema_page(name):
    with open(os.path.join(app.template_folder, "cinema.html"), encoding="utf-8") as f:
        return f.read()

@app.route("/api/refresh", methods=["POST"])
def api_refresh():
    secret = request.headers.get("X-Secret", "")
    if os.environ.get("REFRESH_SECRET") and secret != os.environ["REFRESH_SECRET"]:
        return jsonify({"error": "unauthorized"}), 401
    threading.Thread(target=run_crawl, daemon=True).start()
    return jsonify({"status": "started"})

@app.route("/api/export/excel")
def export_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    df = request.args.get("date_from", date.today().isoformat())
    dt = request.args.get("date_to",   date.today().isoformat())
    cinemas = request.args.getlist("cinema"); mq = request.args.get("movie", "").strip()
    conn = get_db(); cur = conn.cursor()
    sql = "SELECT * FROM screenings WHERE start_dt::date BETWEEN %s AND %s"
    params = [df, dt]
    if cinemas:
        sql += " AND cinema = ANY(%s)"; params.append(cinemas)
    if mq:
        sql += " AND movie ILIKE %s"; params.append(f"%{mq}%")
    cur.execute(sql + " ORDER BY start_dt", params)
    rows = cur.fetchall(); cur.close(); conn.close()

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "시간표"
    headers = ["극장","영화제목","날짜","시작","종료","런타임(분)","상영관","상영유형","프로그램"]
    thin = Side(style="thin", color="DDDDDD"); bdr = Border(left=thin,right=thin,top=thin,bottom=thin)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = PatternFill("solid", fgColor="1A1A2E")
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.alignment = Alignment(horizontal="center"); c.border = bdr
    ws.row_dimensions[1].height = 24
    palette = ["EFF6FF","FFF7ED","F0FDF4","FFF1F2","F5F3FF","ECFEFF","FFFBEB","F0F9FF"]
    colors = {}; pidx = 0
    for ri, r in enumerate(rows, 2):
        if r["cinema"] not in colors:
            colors[r["cinema"]] = palette[pidx % len(palette)]; pidx += 1
        fill = PatternFill("solid", fgColor=colors[r["cinema"]])
        sdt = r["start_dt"] or ""
        for col, v in enumerate([r["cinema"], r["movie"], str(sdt)[:10], str(sdt)[11:16],
                str(r["end_dt"] or "")[11:16], r["runtime"], r["screen"],
                r["show_type"], r["program"]], 1):
            c = ws.cell(row=ri, column=col, value=v)
            c.fill = fill; c.border = bdr; c.alignment = Alignment(vertical="center")
    for col, w in zip("ABCDEFGHI", [16,36,12,10,10,10,14,16,24]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"시간표_{df}_{dt}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ── 시작 ─────────────────────────────────────────────
ensure_db()

def initial_crawl_if_empty():
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("SELECT COUNT(*) AS c FROM screenings")
        count = cur.fetchone()["c"]; cur.close(); conn.close()
        if count == 0:
            log.info("DB 비어있음 → 최초 수집 시작 (백그라운드)")
            threading.Thread(target=run_crawl, daemon=True).start()
        else:
            log.info(f"기존 DB 사용: {count}건")
    except Exception as e:
        log.error(f"DB 초기화 오류: {e}")

initial_crawl_if_empty()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)

# ── 관리자 ────────────────────────────────────────────
@app.route("/admin", methods=["GET", "POST"])
def admin():
    if not ADMIN_PASSWORD:
        return "관리자 페이지가 비활성화되어 있습니다. ADMIN_PASSWORD 환경변수를 설정해주세요.", 503
    ip = _get_ip()
    if request.method == "POST" and "password" in request.form:
        if _is_locked(ip):
            rec = _login_attempts.get(ip, {})
            remain = int((rec["locked_until"] - datetime.now()).total_seconds() // 60) + 1 if rec.get("locked_until") else LOCKOUT_MINUTES
            return render_template_string(ADMIN_LOGIN_TEMPLATE,
                error=f"로그인 시도가 너무 많습니다. {remain}분 후 다시 시도해주세요.")
        if request.form["password"] == ADMIN_PASSWORD:
            _clear_attempts(ip)
            session["admin"] = True
            return redirect(url_for("admin_dashboard"))
        _record_failure(ip)
        rec = _login_attempts.get(ip, {})
        if rec.get("locked_until"):
            error = f"비밀번호가 틀렸습니다. 시도 횟수 초과로 {LOCKOUT_MINUTES}분간 잠금됩니다."
        else:
            left = MAX_ATTEMPTS - rec.get("count", 0)
            error = f"비밀번호가 틀렸습니다. (남은 시도: {left}회)"
        return render_template_string(ADMIN_LOGIN_TEMPLATE, error=error)
    if session.get("admin"):
        return redirect(url_for("admin_dashboard"))
    return render_template_string(ADMIN_LOGIN_TEMPLATE, error=None)


@app.route("/api/events")
def api_events():
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT * FROM events WHERE event_date >= CURRENT_DATE ORDER BY event_date, start_time")
    rows = [dict(r) for r in cur.fetchall()]
    cur.close(); conn.close()
    return jsonify(rows)

@app.route("/api/events", methods=["POST"])
def api_events_post():
    if not session.get("admin"):
        return jsonify({"error": "unauthorized"}), 401
    d = request.get_json()
    conn = get_db(); cur = conn.cursor()
    cur.execute("""
        INSERT INTO events (title, cinema, event_date, start_time, end_time, description, is_free, url)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id
    """, (d["title"], d["cinema"], d["event_date"], d.get("start_time",""),
          d.get("end_time",""), d.get("description",""), bool(d.get("is_free")), d.get("url","")))
    new_id = cur.fetchone()["id"]
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True, "id": new_id})

@app.route("/api/events/<int:eid>", methods=["DELETE"])
def api_events_delete(eid):
    if not session.get("admin"):
        return jsonify({"error": "unauthorized"}), 401
    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM events WHERE id=%s", (eid,))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True})


@app.route("/api/upload-poster", methods=["POST"])
def api_upload_poster():
    if not session.get("admin"):
        return jsonify({"error": "unauthorized"}), 401
    title = request.form.get("title","").strip()
    f = request.files.get("file")
    if not title or not f:
        return jsonify({"error": "title and file required"}), 400
    mime = f.mimetype or "image/jpeg"
    data = f.read()
    if len(data) > 5 * 1024 * 1024:  # 5MB 제한
        return jsonify({"error": "파일이 너무 큽니다 (최대 5MB)"}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute("""
        INSERT INTO posters (title, data, mime, updated_at)
        VALUES (%s, %s, %s, NOW())
        ON CONFLICT (title) DO UPDATE SET
            data=EXCLUDED.data, mime=EXCLUDED.mime, updated_at=NOW()
    """, (title, psycopg2.Binary(data), mime))
    # movies 테이블 poster_url도 업데이트
    poster_url = f"/api/poster/{title}"
    cur.execute("""
        INSERT INTO movies (title, poster_url, updated_at)
        VALUES (%s, %s, NOW())
        ON CONFLICT (title) DO UPDATE SET poster_url=EXCLUDED.poster_url, updated_at=NOW()
    """, (title, poster_url))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True, "poster_url": poster_url})

@app.route("/api/poster/<path:title>")
def api_poster_serve(title):
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT data, mime FROM posters WHERE title=%s", (title,))
    row = cur.fetchone()
    cur.close(); conn.close()
    if not row:
        return "", 404
    from flask import Response
    return Response(bytes(row["data"]), mimetype=row["mime"],
                    headers={"Cache-Control": "public, max-age=86400"})

@app.route("/api/movie-detail")
def api_movie_detail_get():
    title = request.args.get("title","").strip()
    if not title:
        return jsonify({}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT * FROM movies WHERE title=%s", (title,))
    row = cur.fetchone()
    cur.close(); conn.close()
    return jsonify(dict(row) if row else {})

@app.route("/api/movie-detail", methods=["POST"])
def api_movie_detail_post():
    if not session.get("admin"):
        return jsonify({"error": "unauthorized"}), 401
    d = request.get_json()
    title    = (d.get("title") or "").strip()
    if not title:
        return jsonify({"error": "title required"}), 400
    poster   = (d.get("poster_url") or "").strip()
    director = (d.get("director") or "").strip()
    synopsis = (d.get("synopsis") or "").strip()
    year     = int(d["year"]) if d.get("year") else None
    conn = get_db(); cur = conn.cursor()
    cur.execute("""
        INSERT INTO movies (title, poster_url, director, synopsis, year, updated_at)
        VALUES (%s,%s,%s,%s,%s,NOW())
        ON CONFLICT (title) DO UPDATE SET
            poster_url=EXCLUDED.poster_url,
            director=EXCLUDED.director,
            synopsis=EXCLUDED.synopsis,
            year=EXCLUDED.year,
            updated_at=NOW()
    """, (title, poster, director, synopsis, year))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True})

@app.route("/api/movie-info")
def api_movie_info():
    """영화별 추천·수상·시놉시스 통합 반환"""
    conn = get_db(); cur = conn.cursor()
    cur.execute("""
        SELECT m.title, m.director, m.synopsis, m.poster_url,
               r.is_rec, r.comment, r.awards
        FROM movies m
        LEFT JOIN recommended r ON r.title = m.title
    """)
    rows = {r["title"]: dict(r) for r in cur.fetchall()}
    # recommended에만 있는 경우도 포함
    cur.execute("SELECT title, is_rec, comment, awards FROM recommended")
    for r in cur.fetchall():
        if r["title"] not in rows:
            rows[r["title"]] = dict(r)
        else:
            rows[r["title"]].update({k: r[k] for k in ["is_rec","comment","awards"]})
    cur.close(); conn.close()
    return jsonify(rows)

@app.route("/admin/dashboard")
def admin_dashboard():
    if not session.get("admin"):
        return redirect(url_for("admin"))
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT * FROM cinemas ORDER BY name")
    cinemas = [dict(r) for r in cur.fetchall()]
    cur.execute("SELECT * FROM movies ORDER BY title")
    movies = [dict(r) for r in cur.fetchall()]
    # 현재 상영 중인 영화 + 등록된 특별상영 통합 목록
    cur.execute("""
        SELECT DISTINCT sq.movie,
               BOOL_OR(sq.is_event) AS is_event,
               m.poster_url, m.director,
               r.id AS rec_id, r.is_rec, r.comment, r.awards
        FROM (
            SELECT movie, FALSE AS is_event FROM screenings WHERE start_dt::date >= CURRENT_DATE
            UNION ALL
            SELECT title AS movie, TRUE AS is_event FROM events WHERE event_date >= CURRENT_DATE
        ) sq
        LEFT JOIN movies m ON m.title = sq.movie
        LEFT JOIN recommended r ON r.title = sq.movie
        GROUP BY sq.movie, m.poster_url, m.director, r.id, r.is_rec, r.comment, r.awards
        ORDER BY sq.movie
    """)
    screening_movies = [dict(r) for r in cur.fetchall()]
    cur.execute("SELECT * FROM events ORDER BY event_date DESC, start_time")
    events = [dict(r) for r in cur.fetchall()]
    cur.close(); conn.close()
    return render_template_string(ADMIN_DASHBOARD_TEMPLATE,
        cinemas=cinemas, movies=movies,
        screening_movies=screening_movies, events=events)

@app.route("/admin/logout")
def admin_logout():
    session.pop("admin", None)
    return redirect(url_for("admin"))

@app.route("/admin/cinema/save", methods=["POST"])
def admin_cinema_save():
    if not session.get("admin"): return redirect(url_for("admin"))
    d = request.form
    conn = get_db(); cur = conn.cursor()
    cur.execute("""
        INSERT INTO cinemas (name,address,url,phone,description,price_info,price_table,is_free,note,updated_at)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())
        ON CONFLICT (name) DO UPDATE SET
            address=EXCLUDED.address, url=EXCLUDED.url, phone=EXCLUDED.phone,
            description=EXCLUDED.description, price_info=EXCLUDED.price_info,
            price_table=EXCLUDED.price_table,
            is_free=EXCLUDED.is_free, note=EXCLUDED.note, updated_at=NOW()
    """, (d["name"], d.get("address",""), d.get("url",""), d.get("phone",""),
          d.get("description",""), d.get("price_info",""),
          __import__("json").dumps({
              "weekday_matinee": d.get("pt_wd_mat",""),
              "weekday_normal":  d.get("pt_wd_nor",""),
              "weekend_matinee": d.get("pt_we_mat",""),
              "weekend_normal":  d.get("pt_we_nor",""),
              "discount":        d.get("pt_discount",""),
          }),
          d.get("is_free","") == "on", d.get("note","")))
    conn.commit(); cur.close(); conn.close()
    return redirect(url_for("admin_dashboard") + "#cinemas")

@app.route("/admin/movie/save", methods=["POST"])
def admin_movie_save():
    if not session.get("admin"): return redirect(url_for("admin"))
    d = request.form
    conn = get_db(); cur = conn.cursor()
    cur.execute("""
        INSERT INTO movies (title,title_en,director,country,year,runtime,synopsis,poster_url,kobis_url,kofa_url,note,updated_at)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())
        ON CONFLICT (title) DO UPDATE SET
            title_en=EXCLUDED.title_en, director=EXCLUDED.director, country=EXCLUDED.country,
            year=EXCLUDED.year, runtime=EXCLUDED.runtime, synopsis=EXCLUDED.synopsis,
            poster_url=EXCLUDED.poster_url, kobis_url=EXCLUDED.kobis_url,
            kofa_url=EXCLUDED.kofa_url, note=EXCLUDED.note, updated_at=NOW()
    """, (d["title"], d.get("title_en",""), d.get("director",""), d.get("country",""),
          int(d["year"]) if d.get("year") else None,
          int(d["runtime"]) if d.get("runtime") else None,
          d.get("synopsis",""), d.get("poster_url",""), d.get("kobis_url",""),
          d.get("kofa_url",""), d.get("note","")))
    conn.commit(); cur.close(); conn.close()
    return redirect(url_for("admin_dashboard") + "#movies")

@app.route("/admin/movie/delete/<int:mid>", methods=["POST"])
def admin_movie_delete(mid):
    if not session.get("admin"): return redirect(url_for("admin"))
    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM movies WHERE id = %s", (mid,))
    conn.commit(); cur.close(); conn.close()
    return redirect(url_for("admin_dashboard") + "#movies")

# ── 시작 ─────────────────────────────────────────────
ensure_db()

def initial_crawl_if_empty():
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("SELECT COUNT(*) AS c FROM screenings")
        count = cur.fetchone()["c"]; cur.close(); conn.close()
        if count == 0:
            log.info("DB 비어있음 → 최초 수집 시작 (백그라운드)")
            threading.Thread(target=run_crawl, daemon=True).start()
        else:
            log.info(f"기존 DB 사용: {count}건")
    except Exception as e:
        log.error(f"DB 초기화 오류: {e}")

initial_crawl_if_empty()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)

# ── 관리자 HTML 템플릿 ─────────────────────────────────
ADMIN_LOGIN_TEMPLATE = """<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8"><title>관리자 로그인</title>
<link href="https://fonts.googleapis.com/css2?family=Pretendard:wght@400;600;700&display=swap" rel="stylesheet">
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Pretendard',sans-serif;background:#F7F8FA;display:flex;align-items:center;justify-content:center;min-height:100vh}
.card{background:#fff;border:1px solid #E2E5EB;border-radius:12px;padding:40px;width:360px;box-shadow:0 4px 16px rgba(0,0,0,.08)}
.logo{text-align:center;font-size:12px;color:#9CA3AF;margin-bottom:24px;letter-spacing:.04em}
h2{font-size:20px;font-weight:700;margin-bottom:24px;text-align:center}
label{font-size:13px;font-weight:600;color:#374151;display:block;margin-bottom:6px}
input{width:100%;padding:10px 14px;border:1.5px solid #D1D5DB;border-radius:7px;font-size:14px;font-family:inherit;outline:none}
input:focus{border-color:#336600}
.btn{width:100%;margin-top:16px;padding:11px;background:#336600;color:#fff;border:none;border-radius:7px;font-size:14px;font-weight:700;cursor:pointer;font-family:inherit}
.btn:hover{background:#2a5200}
.error{color:#DC2626;font-size:13px;margin-top:10px;text-align:center}
</style>
</head>
<body>
<div class="card">
  <div class="logo">SEOUL // INDIE CINEMA</div>
  <h2>관리자 로그인</h2>
  <form method="POST">
    <label>비밀번호</label>
    <input type="password" name="password" autofocus placeholder="비밀번호 입력">
    <button class="btn" type="submit">로그인</button>
    {% if error %}<div class="error">{{ error }}</div>{% endif %}
  </form>
</div>
</body>
</html>"""

ADMIN_DASHBOARD_TEMPLATE = """<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>어드민 — 오늘 뭐 볼까</title>
<link href="https://fonts.googleapis.com/css2?family=Pretendard:wght@400;500;600;700;800&display=swap" rel="stylesheet">
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Pretendard',sans-serif;background:#F5F4F0;color:#111}
header{background:#111;padding:0 24px;position:sticky;top:0;z-index:100}
.hd{max-width:1100px;margin:0 auto;height:52px;display:flex;align-items:center;justify-content:space-between}
.logo{color:#fff;font-size:13px;font-weight:700;letter-spacing:.05em;text-decoration:none}
.hbtns{display:flex;gap:8px}
.btn{padding:6px 14px;border-radius:6px;border:none;cursor:pointer;font-size:12px;font-weight:600;font-family:inherit;text-decoration:none;display:inline-flex;align-items:center;gap:4px;transition:all .12s}
.btn-light{background:#fff;color:#111}.btn-light:hover{background:#eee}
.btn-primary{background:#2D5A1B;color:#fff}.btn-primary:hover{background:#1e3d12}
.btn-danger{background:#fff;color:#DC2626;border:1.5px solid #FECACA}.btn-danger:hover{background:#FEF2F2}
.btn-sm{padding:3px 9px;font-size:11px}
.tabs{max-width:1100px;margin:24px auto 0;padding:0 24px;display:flex;gap:4px}
.tab{padding:9px 18px;font-size:13px;font-weight:600;cursor:pointer;border-radius:8px 8px 0 0;color:#666;background:#E8E6E0;border:none;font-family:inherit;transition:all .12s}
.tab.active{background:#fff;color:#111}
.main{max-width:1100px;margin:0 auto;padding:0 24px 60px}
.section{display:none}.section.active{display:block}
.panel{background:#fff;border-radius:0 12px 12px 12px;padding:24px 28px;margin-bottom:16px;box-shadow:0 1px 4px rgba(0,0,0,.06)}
.panel+.panel{border-radius:12px}
.panel-title{font-size:14px;font-weight:700;margin-bottom:18px;padding-bottom:10px;border-bottom:1.5px solid #F0F0EE;display:flex;align-items:center;justify-content:space-between}
/* 영화 카드 그리드 */
.movie-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(110px,1fr));gap:8px;margin-bottom:4px}
.m-card{border:2px solid #E8E6E0;border-radius:8px;padding:8px;cursor:pointer;transition:all .12s;background:#FAFAF8;position:relative;text-align:center}
.m-card:hover{border-color:#2D5A1B}
.m-card.selected{border-color:#2D5A1B;background:#EEF5E8;box-shadow:0 0 0 3px rgba(45,90,27,.12)}
.m-card.has-info{border-color:#A3C48A}
.m-card.has-rec{border-style:solid;border-color:#F59E0B}
.m-thumb{width:100%;aspect-ratio:2/3;border-radius:5px;object-fit:cover;background:#E8E6E0;display:flex;align-items:center;justify-content:center;font-size:20px;margin-bottom:5px;overflow:hidden}
.m-thumb img{width:100%;height:100%;object-fit:cover}
.m-name{font-size:10px;font-weight:600;line-height:1.3;word-break:keep-all;color:#111}
.m-badge{position:absolute;top:4px;right:4px;font-size:9px;font-weight:700;padding:1px 4px;border-radius:3px}
.rec-badge{background:#F59E0B;color:#fff}
.info-badge{background:#2D5A1B;color:#fff}
/* 편집 패널 */
.edit-panel{background:#F8F7F4;border:1.5px solid #E0DDD8;border-radius:10px;padding:20px;margin-top:12px;display:none}
.edit-panel.open{display:block}
.ep-title{font-size:15px;font-weight:800;margin-bottom:16px;color:#111;display:flex;align-items:center;gap:10px}
.ep-grid{display:grid;gap:12px}
.ep-2{grid-template-columns:1fr 1fr}
.ep-3{grid-template-columns:1fr 1fr 1fr}
.fg{display:flex;flex-direction:column;gap:4px}
.fg.span2{grid-column:span 2}
.fg.span3{grid-column:1/-1}
label.lbl{font-size:10px;font-weight:700;color:#888;text-transform:uppercase;letter-spacing:.04em}
input,select,textarea{padding:8px 11px;border:1.5px solid #E0DDD8;border-radius:7px;font-size:13px;font-family:inherit;outline:none;background:#fff;width:100%;color:#111}
input:focus,select:focus,textarea:focus{border-color:#2D5A1B;box-shadow:0 0 0 3px rgba(45,90,27,.08)}
textarea{resize:vertical}
/* 포스터 업로드 */
.poster-area{display:grid;grid-template-columns:80px 1fr;gap:12px;align-items:start}
.poster-preview{width:80px;height:120px;border-radius:6px;background:#E8E6E0;display:flex;align-items:center;justify-content:center;font-size:24px;overflow:hidden;flex-shrink:0;border:1.5px solid #E0DDD8}
.poster-preview img{width:100%;height:100%;object-fit:cover}
.poster-inputs{display:flex;flex-direction:column;gap:8px}
.upload-btn-wrap{position:relative;display:inline-flex}
.upload-btn{background:#F0F0EE;border:1.5px dashed #C0BDB8;border-radius:6px;padding:7px 14px;font-size:12px;font-weight:600;cursor:pointer;color:#555;font-family:inherit;transition:all .12s;width:100%;text-align:center}
.upload-btn:hover{border-color:#2D5A1B;color:#2D5A1B}
input[type=file]{position:absolute;inset:0;opacity:0;cursor:pointer;width:100%}
/* 별점 */
.stars-row{display:flex;gap:6px;flex-direction:row-reverse;justify-content:flex-end}
.star-rb{display:none}
.star-lb{font-size:24px;cursor:pointer;color:#DDD;transition:color .08s;line-height:1}
.stars-row:has(.star-rb:checked) .star-rb:checked ~ .star-lb,
.stars-row .star-rb:checked ~ .star-lb,
.star-lb:hover,.star-lb:hover ~ .star-lb{color:#F59E0B}
/* 구분선 */
.ep-section{border-top:1px solid #E8E6E0;margin:16px 0 14px;padding-top:14px}
.ep-section-title{font-size:11px;font-weight:700;color:#888;text-transform:uppercase;letter-spacing:.05em;margin-bottom:10px}
/* 저장바 */
.save-bar{margin-top:16px;display:flex;gap:8px;align-items:center;flex-wrap:wrap}
.msg{font-size:12px;font-weight:600;color:#2D5A1B}
.msg.err{color:#DC2626}
/* 테이블 */
table{width:100%;border-collapse:separate;border-spacing:0;font-size:13px}
thead{background:#F8F7F4}
th{padding:9px 12px;text-align:left;font-weight:700;font-size:11px;color:#888;border-bottom:1.5px solid #E8E6E0;white-space:nowrap}
td{padding:9px 12px;border-bottom:1px solid #F0F0EE;vertical-align:middle}
tr:last-child td{border-bottom:none}
tr:hover td{background:#FAFAF8}
.badge{display:inline-block;font-size:10px;font-weight:700;padding:2px 7px;border-radius:4px}
.badge-free{background:#D1FAE5;color:#065F46}
a{color:#2D5A1B;text-decoration:none}
.price-tbl{width:100%;border-collapse:collapse;font-size:13px;margin-top:4px;}
.price-tbl th,.price-tbl td{border:1.5px solid #E0DDD8;padding:6px 10px;text-align:center;}
.price-tbl th{background:#F5F4F0;font-weight:700;font-size:11px;color:#555;}
.price-tbl-label{font-weight:700;font-size:12px;color:#444;background:#FAFAF8;white-space:nowrap;}
.price-tbl input{border:none;outline:none;text-align:center;font-size:13px;font-family:inherit;width:100%;background:transparent;padding:2px 0;}
/* 극장 카드 */
.cinema-list{display:flex;flex-direction:column;gap:6px;margin-bottom:12px}
.c-row{display:flex;align-items:center;gap:10px;padding:10px 14px;border:1.5px solid #E8E6E0;border-radius:8px;cursor:pointer;background:#FAFAF8;transition:all .12s}
.c-row:hover{border-color:#2D5A1B}
.c-row.selected{border-color:#2D5A1B;background:#EEF5E8}
.c-name{font-size:13px;font-weight:700;flex:1}
.c-desc{font-size:11px;color:#888}
.add-new-btn{background:none;border:1.5px dashed #C0BDB8;border-radius:8px;padding:9px;font-size:12px;font-weight:600;color:#888;cursor:pointer;font-family:inherit;width:100%;transition:all .12s}
.add-new-btn:hover{border-color:#2D5A1B;color:#2D5A1B}
</style>
</head>
<body>
<header>
  <div class="hd">
    <a href="/" class="logo">오늘 뭐 볼까 — 어드민</a>
    <div class="hbtns">
      <a href="/" class="btn btn-light">← 사이트</a>
      <a href="/admin/logout" class="btn btn-light">로그아웃</a>
    </div>
  </div>
</header>

<div class="tabs">
  <button class="tab active" onclick="switchTab('movies')">🎞 영화 관리</button>
  <button class="tab" onclick="switchTab('events')">🎭 특별상영</button>
  <button class="tab" onclick="switchTab('cinemas')">🎬 극장 관리</button>
</div>

<div class="main">

<!-- ══════════════════════════════════════
     🎞 영화 관리
══════════════════════════════════════ -->
<div id="tab-movies" class="section active">
  <div class="panel">
    <div class="panel-title">
      <span>현재 상영 중인 영화 <span style="font-size:12px;font-weight:400;color:#888">— 카드를 눌러 편집하세요</span></span>
      <div style="display:flex;gap:6px;align-items:center;">
        <button class="btn btn-light btn-sm" id="movieSortBtn" onclick="toggleAdminMovieSort()" title="정렬 전환">가 ↑</button>
        <button class="btn btn-light btn-sm" onclick="openNewMovie()">+ 직접 추가</button>
      </div>
    </div>

    <div class="movie-grid" id="movieGrid">
      {% for m in screening_movies %}
      <div class="m-card {% if m.poster_url %}has-info{% endif %} {% if m.rec_id %}has-rec{% endif %}"
           id="mc-{{ loop.index }}"
           data-title="{{ m.movie }}"
           data-poster="{{ (m.poster_url or '')|e }}"
           data-director="{{ (m.director or '')|e }}"
           data-synopsis="{{ (m.synopsis or '')|e }}"
           data-stars="{{ m.stars or 0 }}"
           data-comment="{{ (m.comment or '')|e }}"
           data-awards="{{ (m.awards or '')|e }}"
           data-rec-id="{{ m.rec_id or '' }}"
           onclick="selectMovieCard(this)">
        {% if m.rec_id %}<span class="m-badge rec-badge">★{{ m.stars }}</span>
        {% elif m.poster_url %}<span class="m-badge info-badge">✓</span>{% endif %}
        <div class="m-thumb">
          {% if m.poster_url %}<img src="{{ m.poster_url }}" onerror="this.parentNode.innerHTML='🎬'">
          {% else %}🎬{% endif %}
        </div>
        <div class="m-name">{{ m.movie }}</div>
        {% if m.director %}<div style="font-size:9px;color:#888;margin-top:2px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;">{{ m.director }}</div>{% endif %}
        {% if m.is_event and not m.rec_id %}<span class="m-badge" style="background:#DBEAFE;color:#1E40AF;position:absolute;top:4px;right:4px;font-size:9px;font-weight:700;padding:1px 4px;border-radius:3px;">특별</span>{% endif %}
      </div>
      {% endfor %}
    </div>

    <!-- 편집 패널 -->
    <div class="edit-panel" id="movieEditPanel">
      <div class="ep-title">
        <span id="epMovieTitle">영화 제목</span>
        <button class="btn btn-danger btn-sm" id="epRecDelBtn" onclick="deleteRec()" style="display:none">추천 해제</button>
      </div>
      <input type="hidden" id="epTitle">

      <!-- 포스터 -->
      <div class="ep-section-title">포스터</div>
      <div class="poster-area">
        <div class="poster-preview" id="epPosterPreview">🎬</div>
        <div class="poster-inputs">
          <div class="fg">
            <label class="lbl">URL 직접 입력</label>
            <input type="url" id="epPosterUrl" placeholder="https://..." oninput="onPosterUrlInput(this.value)">
          </div>
          <div class="upload-btn-wrap">
            <button class="upload-btn">📁 파일 업로드</button>
            <input type="file" accept="image/*" onchange="onPosterFileSelect(this)">
          </div>
          <div id="epUploadStatus" style="font-size:11px;color:#888"></div>
        </div>
      </div>

      <!-- 기본 정보 -->
      <div class="ep-section">
        <div class="ep-section-title">기본 정보</div>
        <div class="ep-grid ep-2">
          <div class="fg"><label class="lbl">감독</label>
            <input type="text" id="epDirector" placeholder="예: 봉준호"></div>
          <div class="fg"><label class="lbl">제작연도</label>
            <input type="number" id="epYear" placeholder="2024" min="1900" max="2099"></div>
          <div class="fg span2"><label class="lbl">시놉시스</label>
            <textarea id="epSynopsis" rows="3" placeholder="줄거리를 입력하세요"></textarea></div>
        </div>
      </div>

      <!-- 추천 -->
      <div class="ep-section">
        <div class="ep-section-title">추천</div>
        <div style="display:flex;align-items:center;gap:14px;flex-wrap:wrap;margin-bottom:10px">
          <label style="display:inline-flex;align-items:center;gap:7px;cursor:pointer;font-size:14px;font-weight:700">
            <input type="checkbox" id="epIsRec" style="width:16px;height:16px;accent-color:#111;cursor:pointer"> ★ 추천 영화로 등록
          </label>
          <span style="font-size:11px;color:#888">체크하면 카드에 ★ 표시</span>
        </div>
        <div class="fg">
          <label class="lbl">추천 코멘트 <span style="font-weight:400;text-transform:none">(한 줄, 80자 이내)</span></label>
          <input type="text" id="epComment" placeholder="예: 올해 본 것 중 가장 인상적인 영화" maxlength="80">
        </div>
      </div>

      <!-- 수상 -->
      <div class="ep-section">
        <div class="ep-section-title">수상 내역</div>
        <div class="fg">
          <label class="lbl">쉼표(,)로 구분 &nbsp;예: 2025 칸:황금종려상, 2024 아카데미:각본상 후보</label>
          <textarea id="epAwards" rows="2" placeholder="2025 칸:황금종려상 수상"></textarea>
        </div>
      </div>

      <div class="save-bar">
        <button class="btn btn-primary" onclick="saveMovie()">저장</button>
        <span class="msg" id="epMsg"></span>
      </div>
    </div>
  </div>
</div>

<!-- ══════════════════════════════════════
     🎭 특별상영
══════════════════════════════════════ -->
<div id="tab-events" class="section">
  <div class="panel">
    <div class="panel-title">🎭 특별상영 등록
      <span style="font-size:12px;font-weight:400;color:#888">시사회·GV·특별상영회 등 예매 없는 이벤트</span>
    </div>
    <div class="ep-grid ep-3">
      <div class="fg span2"><label class="lbl">영화 제목 *</label>
        <input type="text" id="evTitle" placeholder="예: 벌새 특별상영"></div>
      <div class="fg"><label class="lbl">극장 *</label>
        <select id="evCinema">
          {% for c in cinemas %}<option>{{ c.name }}</option>{% endfor %}
        </select>
      </div>
      <div class="fg"><label class="lbl">날짜 *</label>
        <input type="date" id="evDate"></div>
      <div class="fg"><label class="lbl">시작</label>
        <input type="time" id="evStart"></div>
      <div class="fg"><label class="lbl">종료</label>
        <input type="time" id="evEnd"></div>
      <div class="fg span3"><label class="lbl">설명</label>
        <textarea id="evDesc" rows="2" placeholder="감독 GV 포함. 당일 선착순 입장"></textarea></div>
      <div class="fg"><label class="lbl">링크</label>
        <input type="url" id="evUrl" placeholder="https://..."></div>
      <div class="fg" style="justify-content:flex-end;padding-top:18px">
        <label style="display:inline-flex;align-items:center;gap:6px;cursor:pointer;font-size:13px;font-weight:600">
          <input type="checkbox" id="evFree" style="width:auto"> 무료
        </label>
      </div>
    </div>
    <div class="save-bar">
      <button class="btn btn-primary" onclick="saveEvent()">등록</button>
      <span class="msg" id="evMsg"></span>
    </div>
  </div>
  <div class="panel">
    <div class="panel-title">등록된 특별상영 ({{ events|length }}건)</div>
    {% if events %}
    <div style="overflow-x:auto"><table>
      <thead><tr><th>날짜</th><th>극장</th><th>제목</th><th>시간</th><th>설명</th><th>무료</th><th></th></tr></thead>
      <tbody>{% for e in events %}
      <tr id="ev-{{ e.id }}">
        <td style="white-space:nowrap">{{ e.event_date }}</td>
        <td>{{ e.cinema }}</td>
        <td><strong>{{ e.title }}</strong>{% if e.url %} <a href="{{ e.url }}" target="_blank">↗</a>{% endif %}</td>
        <td style="white-space:nowrap">{{ e.start_time or "—" }}{% if e.end_time %} – {{ e.end_time }}{% endif %}</td>
        <td style="font-size:12px;color:#555;max-width:180px">{{ e.description or "—" }}</td>
        <td>{% if e.is_free %}<span class="badge badge-free">무료</span>{% else %}—{% endif %}</td>
        <td><button class="btn btn-danger btn-sm" onclick="deleteEvent({{ e.id }})">삭제</button></td>
      </tr>{% endfor %}</tbody>
    </table></div>
    {% else %}<p style="color:#aaa;font-size:13px">등록된 특별상영이 없습니다.</p>{% endif %}
  </div>
</div>

<!-- ══════════════════════════════════════
     🎬 극장 관리
══════════════════════════════════════ -->
<div id="tab-cinemas" class="section">
  <div class="panel">
    <div class="panel-title">
      <span>극장 목록 <span style="font-size:12px;font-weight:400;color:#888">— 극장을 눌러 수정하세요</span></span>
      <div style="display:flex;gap:6px;align-items:center;">
        <button class="btn btn-light btn-sm" id="cinemaSortBtn" onclick="toggleAdminCinemaSort()" title="정렬 전환">가 ↑</button>
        <button class="btn btn-light btn-sm" onclick="openNewCinema()">+ 신규 추가</button>
      </div>
    </div>
    <div class="cinema-list">
      {% for c in cinemas %}
      <div class="c-row" onclick="selectCinema(this)"
           data-name="{{ c.name }}"
           data-address="{{ (c.address or '')|e }}"
           data-url="{{ (c.url or '')|e }}"
           data-phone="{{ (c.phone or '')|e }}"
           data-description="{{ (c.description or '')|e }}"
           data-price="{{ (c.price_info or '')|e }}"
           data-pricetable="{{ (c.price_table|tojson if c.price_table else '{}')|e }}"
           data-note="{{ (c.note or '')|e }}"
           data-free="{{ 'true' if c.is_free else 'false' }}">
        <div>
          <div class="c-name">{{ c.name }}</div>
          <div class="c-desc">{{ c.address or "" }}{% if c.price_info %} · {{ c.price_info[:30] }}{% if c.price_info|length > 30 %}…{% endif %}{% endif %}</div>
        </div>
        {% if c.is_free %}<span class="badge badge-free">무료</span>{% endif %}
      </div>
      {% endfor %}
    </div>

    <div class="edit-panel" id="cinemaEditPanel">
      <div class="ep-title" id="epCinemaTitle">극장명</div>
      <form method="POST" action="/admin/cinema/save" id="cinemaForm">
        <div class="ep-grid ep-2">
          <div class="fg"><label class="lbl">극장명 *</label>
            <input type="text" name="name" id="ciName" required></div>
          <div class="fg"><label class="lbl">홈페이지</label>
            <input type="url" name="url" id="ciUrl" placeholder="https://..."></div>
          <div class="fg"><label class="lbl">주소</label>
            <input type="text" name="address" id="ciAddress"></div>
          <div class="fg"><label class="lbl">전화</label>
            <input type="text" name="phone" id="ciPhone" placeholder="02-..."></div>
          <input type="hidden" name="price_info" id="ciPrice">
          <div class="fg span2"><label class="lbl">특징</label>
            <textarea name="description" id="ciDesc" rows="2"></textarea></div>
          <!-- 요금표 2×2 -->
          <div class="fg span2">
            <label class="lbl">요금표</label>
            <table class="price-tbl">
              <thead><tr><th></th><th>조조</th><th>일반</th></tr></thead>
              <tbody>
                <tr>
                  <td class="price-tbl-label">월~목</td>
                  <td><input type="text" name="pt_wd_mat" id="ptWdMat" placeholder="예: 7,000원" style="width:100%"></td>
                  <td><input type="text" name="pt_wd_nor" id="ptWdNor" placeholder="예: 11,000원" style="width:100%"></td>
                </tr>
                <tr>
                  <td class="price-tbl-label">금~일·공휴일</td>
                  <td><input type="text" name="pt_we_mat" id="ptWeMat" placeholder="예: 8,000원" style="width:100%"></td>
                  <td><input type="text" name="pt_we_nor" id="ptWeNor" placeholder="예: 12,000원" style="width:100%"></td>
                </tr>
              </tbody>
            </table>
          </div>
          <div class="fg span2">
            <label class="lbl">할인·기타 사항 <span style="font-weight:400;text-transform:none">— 예: 청소년 1,000원 할인 / 회원 2,000원 할인</span></label>
            <textarea name="pt_discount" id="ptDiscount" rows="2" placeholder="예: 청소년 1,000원 할인 / 멤버십 회원 2,000원 할인 / 조조 기준: 첫 회차"></textarea>
          </div>
          <div class="fg"><label class="lbl">메모 (내부용)</label>
            <input type="text" name="note" id="ciNote"></div>
          <div class="fg" style="justify-content:flex-end;padding-top:18px">
            <label style="display:inline-flex;align-items:center;gap:6px;cursor:pointer;font-size:13px;font-weight:600">
              <input type="checkbox" name="is_free" id="ciFree" style="width:auto"> 무료 상영관
            </label>
          </div>
        </div>
        <div class="save-bar">
          <button type="submit" class="btn btn-primary">저장</button>
        </div>
      </form>
    </div>
  </div>
</div>

</div><!-- /main -->

<script>
const TABS = ["movies","events","cinemas"];
function switchTab(name){
  document.querySelectorAll(".tab").forEach((t,i)=>t.classList.toggle("active",TABS[i]===name));
  document.querySelectorAll(".section").forEach(s=>s.classList.remove("active"));
  document.getElementById("tab-"+name).classList.add("active");
}

// ── 영화 카드 선택 ────────────────────
let selCard = null;

function selectMovieCard(card){
  if(selCard===card){
    card.classList.remove("selected");
    selCard=null;
    document.getElementById("movieEditPanel").classList.remove("open");
    return;
  }
  document.querySelectorAll(".m-card").forEach(c=>c.classList.remove("selected"));
  card.classList.add("selected");
  selCard=card;
  fillMoviePanel(card);
  const panel=document.getElementById("movieEditPanel");
  panel.classList.add("open");
  setTimeout(()=>panel.scrollIntoView({behavior:"smooth",block:"nearest"}),50);
}

function fillMoviePanel(card){
  const d=card.dataset;
  document.getElementById("epTitle").value    = d.title;
  document.getElementById("epMovieTitle").textContent = d.title;
  document.getElementById("epPosterUrl").value= d.poster||"";
  document.getElementById("epDirector").value = d.director||"";
  document.getElementById("epYear").value     = d.year||"";
  document.getElementById("epSynopsis").value = d.synopsis||"";
  document.getElementById("epComment").value  = d.comment||"";
  document.getElementById("epAwards").value   = d.awards||"";
  document.getElementById("epMsg").textContent= "";
  document.getElementById("epUploadStatus").textContent="";

  // 포스터 미리보기
  setPosterPreview(d.poster||"");

  // 추천 체크박스
  document.getElementById("epIsRec").checked = !!d.recId;
  const delBtn=document.getElementById("epRecDelBtn");
  delBtn.style.display = d.recId ? "inline-flex":"none";
  delBtn.dataset.recId = d.recId||"";
}

function openNewMovie(){
  document.querySelectorAll(".m-card").forEach(c=>c.classList.remove("selected"));
  selCard=null;
  // 빈 패널 열기
  document.getElementById("epTitle").value="";
  document.getElementById("epMovieTitle").textContent="새 영화 추가";
  ["epPosterUrl","epDirector","epYear","epSynopsis","epComment","epAwards"].forEach(id=>document.getElementById(id).value="");
  document.getElementById("epIsRec").checked=false;
  document.getElementById("epRecDelBtn").style.display="none";
  document.querySelectorAll("input[name='epStars']").forEach(r=>r.checked=false);
  setPosterPreview("");
  document.getElementById("epMsg").textContent="";
  const panel=document.getElementById("movieEditPanel");
  panel.classList.add("open");
  panel.scrollIntoView({behavior:"smooth",block:"start"});
  document.getElementById("epTitle").focus();
  // 제목 입력 가능하게 교체
  const titleEl=document.getElementById("epMovieTitle");
  const input=document.createElement("input");
  input.type="text"; input.placeholder="영화 제목 입력";
  input.style.cssText="font-size:15px;font-weight:800;border:none;border-bottom:2px solid #2D5A1B;background:transparent;outline:none;width:200px;padding:0";
  input.id="epTitleInput";
  titleEl.replaceWith(input);
  input.focus();
}

// ── 포스터 ───────────────────────────
function setPosterPreview(url){
  const el=document.getElementById("epPosterPreview");
  el.innerHTML = url ? `<img src="${url}" style="width:100%;height:100%;object-fit:cover" onerror="this.parentNode.innerHTML='🎬'">` : "🎬";
}
function onPosterUrlInput(url){
  setPosterPreview(url);
}

async function onPosterFileSelect(input){
  const file=input.files[0]; if(!file) return;
  const title=document.getElementById("epTitle").value || (document.getElementById("epTitleInput")?.value||"");
  if(!title){ alert("영화 제목을 먼저 입력하세요."); return; }

  document.getElementById("epUploadStatus").textContent="업로드 중…";
  const fd=new FormData();
  fd.append("title",title); fd.append("file",file);

  // 미리보기 즉시
  const reader=new FileReader();
  reader.onload=e=>setPosterPreview(e.target.result);
  reader.readAsDataURL(file);

  const res=await fetch("/api/upload-poster",{method:"POST",body:fd});
  if(res.ok){
    const d=await res.json();
    document.getElementById("epPosterUrl").value=d.poster_url;
    document.getElementById("epUploadStatus").textContent="✓ 업로드 완료";
    if(selCard) selCard.dataset.poster=d.poster_url;
  } else {
    document.getElementById("epUploadStatus").textContent="❌ 업로드 실패";
  }
}

// ── 영화 저장 ─────────────────────────
async function saveMovie(){
  // 새 영화 추가 모드면 title 인풋에서 가져오기
  const titleInput=document.getElementById("epTitleInput");
  const title = titleInput ? titleInput.value.trim() : document.getElementById("epTitle").value.trim();
  if(!title){ alert("영화 제목을 입력하세요."); return; }

  const poster  =document.getElementById("epPosterUrl").value.trim();
  const director=document.getElementById("epDirector").value.trim();
  const year    =document.getElementById("epYear").value;
  const synopsis=document.getElementById("epSynopsis").value.trim();
  const comment =document.getElementById("epComment").value.trim();
  const awards  =document.getElementById("epAwards").value.trim();
  const isRec   =document.getElementById("epIsRec").checked;
  // 1) 기본 정보 저장
  const r1=await fetch("/api/movie-detail",{
    method:"POST",headers:{"Content-Type":"application/json"},
    body:JSON.stringify({title,poster_url:poster,director,year:year||null,synopsis})
  });

  // 2) 추천 저장 or 해제
  if(isRec){
    await fetch("/api/recommended",{
      method:"POST",headers:{"Content-Type":"application/json"},
      body:JSON.stringify({title,is_rec:true,comment,awards})
    });
  } else if(selCard?.dataset.recId){
    await fetch(`/api/recommended/${selCard.dataset.recId}`,{method:"DELETE"});
  } else if(awards){
    // 추천 없이 수상만
    await fetch("/api/recommended",{
      method:"POST",headers:{"Content-Type":"application/json"},
      body:JSON.stringify({title,stars:0,comment:"",awards})
    });
  }

  if(r1.ok){
    const msg=document.getElementById("epMsg");
    msg.textContent="✓ 저장됐어요"; msg.className="msg";
    // 카드 업데이트
    if(selCard){
      selCard.dataset.director=director;
      selCard.dataset.synopsis=synopsis;
      selCard.dataset.poster=poster;
      selCard.dataset.comment=comment;
      selCard.dataset.awards=awards;
      if(poster){
        selCard.classList.add("has-info");
        selCard.querySelector(".m-thumb").innerHTML=`<img src="${poster}" style="width:100%;height:100%;object-fit:cover" onerror="this.parentNode.innerHTML='🎬'">`;
      }
      if(isRec){
        selCard.classList.add("has-rec");
        let badge=selCard.querySelector(".m-badge.rec-badge");
        if(!badge){ badge=document.createElement("span"); badge.className="m-badge rec-badge"; selCard.appendChild(badge); }
        badge.textContent="★";
      }
    }
    setTimeout(()=>msg.textContent="",3000);
  } else {
    const msg=document.getElementById("epMsg");
    msg.textContent="❌ 저장 실패"; msg.className="msg err";
  }
}

async function deleteRec(){
  if(!confirm("추천을 해제할까요?")) return;
  const id=document.getElementById("epRecDelBtn").dataset.recId;
  if(!id) return;
  await fetch(`/api/recommended/${id}`,{method:"DELETE"});
  document.getElementById("epIsRec").checked=false;
  document.getElementById("epRecDelBtn").style.display="none";
  if(selCard){ selCard.classList.remove("has-rec"); selCard.dataset.recId=""; }
}

// ── 특별상영 ──────────────────────────
async function saveEvent(){
  const title=document.getElementById("evTitle").value.trim();
  const cinema=document.getElementById("evCinema").value;
  const date=document.getElementById("evDate").value;
  if(!title||!cinema||!date){alert("제목, 극장, 날짜는 필수예요.");return;}
  const res=await fetch("/api/events",{
    method:"POST",headers:{"Content-Type":"application/json"},
    body:JSON.stringify({title,cinema,event_date:date,
      start_time:document.getElementById("evStart").value,
      end_time:document.getElementById("evEnd").value,
      description:document.getElementById("evDesc").value.trim(),
      is_free:document.getElementById("evFree").checked,
      url:document.getElementById("evUrl").value.trim()})
  });
  if(res.ok){
    document.getElementById("evMsg").textContent="✓ 등록됐어요";
    setTimeout(()=>location.reload(),800);
  } else { document.getElementById("evMsg").textContent="❌ 등록 실패"; }
}
async function deleteEvent(id){
  if(!confirm("삭제할까요?"))return;
  await fetch(`/api/events/${id}`,{method:"DELETE"});
  document.getElementById("ev-"+id)?.remove();
}

// ── 극장 ──────────────────────────────
function selectCinema(row){
  document.querySelectorAll(".c-row").forEach(r=>r.classList.remove("selected"));
  row.classList.add("selected");
  const d=row.dataset;
  document.getElementById("epCinemaTitle").textContent=d.name;
  document.getElementById("ciName").value   =d.name;
  document.getElementById("ciUrl").value    =d.url;
  document.getElementById("ciAddress").value=d.address;
  document.getElementById("ciPhone").value  =d.phone;
  document.getElementById("ciDesc").value   =d.description;
  document.getElementById("ciPrice").value  =d.price;
  // 요금표
  let pt={};
  try{ pt=JSON.parse(d.pricetable||"{}"); }catch(e){}
  document.getElementById("ptWdMat").value  =pt.weekday_matinee||"";
  document.getElementById("ptWdNor").value  =pt.weekday_normal||"";
  document.getElementById("ptWeMat").value  =pt.weekend_matinee||"";
  document.getElementById("ptWeNor").value  =pt.weekend_normal||"";
  document.getElementById("ptDiscount").value=pt.discount||"";
  document.getElementById("ciNote").value   =d.note;
  document.getElementById("ciFree").checked =(d.free==="true");
  const panel=document.getElementById("cinemaEditPanel");
  panel.classList.add("open");
  panel.scrollIntoView({behavior:"smooth",block:"nearest"});
}
function openNewCinema(){
  document.querySelectorAll(".c-row").forEach(r=>r.classList.remove("selected"));
  ["ciName","ciUrl","ciAddress","ciPhone","ciDesc","ciPrice","ciNote","ptWdMat","ptWdNor","ptWeMat","ptWeNor","ptDiscount"].forEach(id=>document.getElementById(id).value="");
  document.getElementById("ciFree").checked=false;
  document.getElementById("epCinemaTitle").textContent="새 극장 추가";
  const panel=document.getElementById("cinemaEditPanel");
  panel.classList.add("open");
  panel.scrollIntoView({behavior:"smooth",block:"nearest"});
  document.getElementById("ciName").focus();
}

// 오늘 날짜 기본값
document.getElementById("evDate").value=new Date().toISOString().slice(0,10);
/* ── 어드민 정렬 ── */
let adminMovieSortAsc = true;
let adminCinemaSortAsc = true;

function toggleAdminMovieSort(){
  adminMovieSortAsc = !adminMovieSortAsc;
  const btn = document.getElementById("movieSortBtn");
  btn.textContent = adminMovieSortAsc ? "가 ↑" : "가 ↓";
  const grid = document.getElementById("movieGrid");
  const cards = [...grid.querySelectorAll(".m-card")];
  cards.sort((a,b)=>{
    const ta = (a.dataset.title||"").localeCompare(b.dataset.title||"","ko");
    return adminMovieSortAsc ? ta : -ta;
  });
  cards.forEach(c=>grid.appendChild(c));
}

function toggleAdminCinemaSort(){
  adminCinemaSortAsc = !adminCinemaSortAsc;
  const btn = document.getElementById("cinemaSortBtn");
  btn.textContent = adminCinemaSortAsc ? "가 ↑" : "가 ↓";
  const list = document.querySelector(".cinema-list");
  const rows = [...list.querySelectorAll(".c-row")];
  rows.sort((a,b)=>{
    const ta = (a.dataset.name||"").localeCompare(b.dataset.name||"","ko");
    return adminCinemaSortAsc ? ta : -ta;
  });
  rows.forEach(r=>list.appendChild(r));
}
</script>
</body>
</html>"""
