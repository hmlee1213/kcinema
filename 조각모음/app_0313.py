# app.py — 서울 독립극장 시간표 서비스 (PostgreSQL)
import io, os, threading, logging, time, re
from datetime import datetime, date, timedelta
from flask import Flask, jsonify, request, send_file, session, redirect, url_for, render_template_string
import requests
from bs4 import BeautifulSoup
import psycopg2
from psycopg2.extras import RealDictCursor

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

app = Flask(__name__, template_folder="templates")

# SECRET_KEY: Railway 고유값 조합 자동생성 (별도 설정 불필요)
# 직접 설정한 값이 있으면 그걸 우선 사용
app.secret_key = os.environ.get("SECRET_KEY") or (
    os.environ.get("RAILWAY_PROJECT_ID", "local") + "-" +
    os.environ.get("RAILWAY_SERVICE_ID", "dev")
)

# ADMIN_PASSWORD: Railway Variables에서 반드시 직접 설정
# 설정 안 하면 어드민 페이지 자체가 비활성화됨
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "")

DATABASE_URL = os.environ.get("DATABASE_URL")

# ── 브루트포스 방어 (IP별 시도 횟수, 메모리 저장) ───────
# { ip: {"count": int, "locked_until": datetime | None} }
_login_attempts: dict = {}
MAX_ATTEMPTS    = 10
LOCKOUT_MINUTES = 30

def _get_ip():
    return request.headers.get("X-Forwarded-For", request.remote_addr or "").split(",")[0].strip()

def _is_locked(ip):
    rec = _login_attempts.get(ip)
    if not rec: return False
    if rec["locked_until"] and datetime.now() < rec["locked_until"]:
        return True
    if rec["locked_until"] and datetime.now() >= rec["locked_until"]:
        # 잠금 해제 — 카운트 초기화
        _login_attempts.pop(ip, None)
    return False

def _record_failure(ip):
    rec = _login_attempts.setdefault(ip, {"count": 0, "locked_until": None})
    rec["count"] += 1
    if rec["count"] >= MAX_ATTEMPTS:
        rec["locked_until"] = datetime.now() + timedelta(minutes=LOCKOUT_MINUTES)
        log.warning(f"관리자 로그인 잠금: {ip} ({MAX_ATTEMPTS}회 실패)")

def _clear_attempts(ip):
    _login_attempts.pop(ip, None)

CINEMA_ORDER = [
    "KT&G상상마당시네마","KU시네마테크","더숲아트시네마",
    "라이카시네마","서울아트시네마","씨네큐브",
    "아리랑시네센터","아트나인","아트하우스모모","에무시네마",
    "서울영화센터","한국영상자료원"
]
FREE_CINEMAS = ["서울영화센터","한국영상자료원"]

CINEMAS = [
    {"name":"KU시네마테크","source":"moviee","t_id":"121"},
    {"name":"KT&G상상마당시네마","source":"moviee","t_id":"123"},
    {"name":"서울아트시네마","source":"seoulart"},
    {"name":"한국영상자료원","source":"kofa"},
    {"name":"라이카시네마","source":"dtryx","brand":"spacedog","cinema_cd":"000072"},
    {"name":"씨네큐브","source":"dtryx","brand":"cinecube","cinema_cd":"000003"},
    {"name":"더숲아트시네마","source":"dtryx","brand":"indieart","cinema_cd":"000065"},
    {"name":"아트하우스모모","source":"dtryx","brand":"indieart","cinema_cd":"000067"},
    {"name":"서울영화센터","source":"dtryx","brand":"seoulcc","cinema_cd":"000160"},
    {"name":"아리랑시네센터","source":"dtryx","brand":"etc","cinema_cd":"000088"},
    {"name":"에무시네마","source":"dtryx","brand":"indieart","cinema_cd":"000069"},
    {"name":"아트나인","source":"dtryx","brand":"etc","cinema_cd":"000162"},
]

DTRYX_CGID = "FE8EF4D2-F22D-4802-A39A-D58F23A29C1E"

# ── DB ───────────────────────────────────────────────
def get_db():
    conn = psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)
    return conn

def ensure_db():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS screenings (
            cinema TEXT, movie TEXT, start_dt TEXT, end_dt TEXT,
            runtime INTEGER, screen TEXT, source TEXT,
            show_type TEXT, program TEXT, movie_url TEXT,
            PRIMARY KEY(cinema, start_dt, screen)
        )
    """)
    cur.execute("ALTER TABLE screenings ADD COLUMN IF NOT EXISTS movie_url TEXT")
    cur.execute("ALTER TABLE screenings ADD COLUMN IF NOT EXISTS movie_cd TEXT")
    cur.execute("""
        CREATE TABLE IF NOT EXISTS meta (key TEXT PRIMARY KEY, value TEXT)
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS cinemas (
            name TEXT PRIMARY KEY,
            address TEXT DEFAULT '',
            url TEXT DEFAULT '',
            phone TEXT DEFAULT '',
            description TEXT DEFAULT '',
            is_free BOOLEAN DEFAULT FALSE,
            note TEXT DEFAULT '',
            updated_at TIMESTAMP DEFAULT NOW()
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS movies (
            id SERIAL PRIMARY KEY,
            title TEXT UNIQUE NOT NULL,
            title_en TEXT DEFAULT '',
            director TEXT DEFAULT '',
            country TEXT DEFAULT '',
            year INTEGER,
            runtime INTEGER,
            synopsis TEXT DEFAULT '',
            poster_url TEXT DEFAULT '',
            kobis_url TEXT DEFAULT '',
            kofa_url TEXT DEFAULT '',
            note TEXT DEFAULT '',
            updated_at TIMESTAMP DEFAULT NOW()
        )
    """)
    # 극장 기본 정보 초기 삽입 (없는 경우만)
    cinema_data = [
        ("KU시네마테크","서울 광진구 능동로 120 건국대학교 예술문화관 지하","https://www.kucine.com","02-450-3551","건국대학교 내 시네마테크. 독립·예술영화 및 특별전 중심.",False,""),
        ("KT&G상상마당시네마","서울 마포구 어울마당로 65","https://www.sangsangmadang.com","02-330-6220","홍대 KT&G상상마당 내 위치. 독립·예술·기획 상영 전문.",False,""),
        ("서울아트시네마","서울 종로구 삼일대로 428 낙원상가 4층","https://www.cinematheque.seoul.kr","02-741-9782","한국 최초 시네마테크 전문관. 고전·회고전·특별전 중심.",False,""),
        ("한국영상자료원","서울 마포구 월드컵북로 400","https://www.koreafilm.or.kr","02-3153-2001","한국 영화문화유산 보존기관. 무료 상영 및 아카이브 특별전.",True,"무료 상영"),
        ("라이카시네마","서울 마포구 와우산로29나길 20","https://www.likacinema.com","","홍대 인근 소규모 독립영화 전문관.",False,""),
        ("씨네큐브","서울 종로구 광화문 KT빌딩 지하 2층","https://www.cinecube.co.kr","02-2002-7782","광화문 위치. 예술·독립영화 전문 상영관.",False,""),
        ("더숲아트시네마","서울 강남구 언주로 154","https://www.thesoop.com","","강남 예술영화 전문관. 상영 시작 10분 지연.",False,"정시 기준 10분 지연 상영"),
        ("아트하우스모모","서울 서대문구 신촌역로 22 메가박스 신촌점 5층","https://www.arthousemomo.com","02-332-5558","신촌 메가박스 내 독립·예술영화 전문관.",False,""),
        ("서울영화센터","서울 중구 퇴계로 213","https://www.sfc.kr","02-2285-0562","서울시 지원 독립·예술영화 전용관. 무료 상영.",True,"무료 상영"),
        ("아리랑시네센터","서울 성북구 아리랑로 82","https://www.arirancine.co.kr","02-936-5008","성북구 위치 지역 예술·독립영화관.",False,""),
        ("에무시네마","서울 종로구 창경궁로 89 EMMU 지하","https://www.emucinema.com","","종로 위치 소규모 독립영화 전문관.",False,""),
        ("아트나인","서울 강남구 선릉로 221 강남 S타워 지하","https://www.artnine.co.kr","02-563-5670","강남 예술영화 전문 상영관.",False,""),
    ]
    for c in cinema_data:
        cur.execute(
            "INSERT INTO cinemas (name,address,url,phone,description,is_free,note) VALUES (%s,%s,%s,%s,%s,%s,%s) ON CONFLICT (name) DO NOTHING",
            c
        )
    cur.execute("""
        CREATE TABLE IF NOT EXISTS recommended (
            id SERIAL PRIMARY KEY,
            title TEXT UNIQUE NOT NULL,
            description TEXT DEFAULT '',
            awards TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT NOW()
        )
    """)
    # 기존 테이블 마이그레이션 (reason 컬럼이 있을 경우 대비)
    cur.execute("ALTER TABLE recommended ADD COLUMN IF NOT EXISTS description TEXT DEFAULT ''")
    cur.execute("ALTER TABLE recommended ADD COLUMN IF NOT EXISTS awards TEXT DEFAULT ''")
    conn.commit(); cur.close(); conn.close()

def save_to_db(rows):
    if not rows: return
    conn = get_db(); cur = conn.cursor()
    for r in rows:
        cur.execute("""
            INSERT INTO screenings (cinema,movie,start_dt,end_dt,runtime,screen,source,show_type,program)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT (cinema, start_dt, screen) DO UPDATE SET
                movie=EXCLUDED.movie, end_dt=EXCLUDED.end_dt,
                runtime=EXCLUDED.runtime, source=EXCLUDED.source,
                show_type=EXCLUDED.show_type, program=EXCLUDED.program
        """, (
            r["cinema"], r["movie"],
            str(r["start_dt"]) if r["start_dt"] else None,
            str(r["end_dt"])   if r["end_dt"]   else None,
            r["runtime"], r["screen"], r["source"],
            r.get("show_type",""), r.get("program","")
        ))
    cur.execute("""
        INSERT INTO meta (key, value) VALUES ('last_updated', %s)
        ON CONFLICT (key) DO UPDATE SET value=EXCLUDED.value
    """, (datetime.now().isoformat(),))
    conn.commit(); cur.close(); conn.close()

# ── 공통 함수 (getdb.py 원본) ─────────────────────────
def make_datetime(date_obj, time_str):
    if not time_str: return None
    try:
        return datetime.combine(date_obj, datetime.strptime(time_str,"%H:%M").time())
    except:
        return None

def compute_end_dt(start_dt, runtime):
    return start_dt+timedelta(minutes=runtime) if start_dt and runtime else None

def compute_runtime(start_time, end_time):
    try:
        s = datetime.strptime(start_time, "%H:%M")
        e = datetime.strptime(end_time, "%H:%M")
        return int((e-s).total_seconds()//60)
    except:
        return None

# ── fetch 함수들 (getdb.py 원본 그대로) ──────────────
def fetch_dtryx(cinema, start_date, days=14):
    rows=[]
    for i in range(days):
        day=start_date+timedelta(days=i)
        day_str=day.strftime("%Y-%m-%d")
        params={"cgid":DTRYX_CGID,"BrandCd":cinema["brand"],"CinemaCd":cinema["cinema_cd"],
                "PlaySDT":day_str,"_":int(time.time()*1000)}
        headers={"User-Agent":"Mozilla/5.0","X-Requested-With":"XMLHttpRequest"}
        try:
            data=requests.get("https://www.dtryx.com/cinema/showseq_list.do", params=params, headers=headers, timeout=10).json()
        except:
            continue
        for item in data.get("Showseqlist",[]):
            start=item.get("StartTime"); end=item.get("EndTime")
            runtime=compute_runtime(start,end)
            start_dt=make_datetime(day,start)
            end_dt=make_datetime(day,end) if end else compute_end_dt(start_dt,runtime)
            special_fields=[item.get(f,"").strip() for f in ["ScreenTypeNmNat","PlayTimeTypeNm","DisplayTypeDetailNm","ScreeningInfoNat"]]
            show_type="일반" if all(f in ["","일반"] for f in special_fields) else " / ".join(f for f in special_fields if f not in ["","일반"])
            screen_name=item.get("ScreenNmNat") or item.get("ScreenNm") or ""
            program=item.get("ProgramName","").strip() if item.get("ProgramName") else ""
            rows.append({"cinema":cinema["name"],"movie":item.get("MovieNmNat"),
                         "start_dt":start_dt,"end_dt":end_dt,"runtime":runtime,
                         "screen":screen_name,"source":"dtryx",
                         "show_type":show_type,"program":program})
    return rows

def fetch_moviee(cinema,start_date,days=14):
    rows=[]
    url="https://moviee.co.kr/api/TicketApi/GetPlayTimeList"
    for i in range(days):
        day=start_date+timedelta(days=i)
        params={"tId":cinema["t_id"],"playDt":day.strftime("%Y-%m-%d")}
        try:
            data=requests.get(url, params=params, timeout=10).json()
        except:
            continue
        if data.get("ResCd")!="00": continue
        for item in data["ResData"]["Table"]:
            start=item["PLAY_TIME"]; end=item["END_TIME"]
            start=f"{start[:2]}:{start[2:]}" if start else None
            end=f"{end[:2]}:{end[2:]}" if end else None
            runtime=compute_runtime(start,end)
            start_dt=make_datetime(day,start)
            end_dt=make_datetime(day,end) if end else compute_end_dt(start_dt,runtime)
            movie_name=item["M_NM"]
            show_type=""
            if "(" in movie_name and ")" in movie_name:
                main_title=movie_name.split("(")[0].strip()
                show_type=movie_name.split("(")[1].replace(")","").strip()
            else:
                main_title=movie_name.strip()
            screen_name=item.get("ROOM_NM","")
            rows.append({"cinema":cinema["name"],"movie":main_title,
                         "start_dt":start_dt,"end_dt":end_dt,"runtime":runtime,
                         "screen":screen_name,"source":"moviee",
                         "show_type":show_type,"program":""})
    return rows

def fetch_seoulart(cinema):
    rows=[]
    url="https://www.cinematheque.seoul.kr/bbs/content.php?co_id=timetable"
    try:
        res=requests.get(url, timeout=10)
        soup=BeautifulSoup(res.text,"html.parser")
    except:
        return rows
    year=datetime.today().year
    theater_name=cinema["name"]
    tables=soup.select("table")
    for table in tables:
        date_rows=table.select("tr.date-label")
        for dr in date_rows:
            tds=dr.find_all("td")
            col_dates=[]
            for td in tds:
                txt=td.get_text(strip=True)
                if txt:
                    m,d,_=txt.split(".")
                    col_dates.append(datetime(year,int(m),int(d)))
                else:
                    col_dates.append(None)
            next_tr=dr.find_next_sibling()
            while next_tr and "event" in next_tr.get("class",[]):
                for idx, td in enumerate(next_tr.find_all("td")):
                    date_obj=col_dates[idx] if idx<len(col_dates) else None
                    if not date_obj: continue
                    for link in td.find_all("a"):
                        try:
                            start=link.find("strong").text.strip()
                            title_tag=link.find_all("p")[1]
                            title=re.sub(r"\(\d+min\)","",title_tag.text).strip()
                            runtime_match=re.search(r"\((\d+)min\)",title_tag.text)
                            runtime=int(runtime_match.group(1)) if runtime_match else None
                            start_dt=make_datetime(date_obj,start)
                            end_dt=compute_end_dt(start_dt,runtime)
                            rows.append({"cinema":theater_name,"movie":title,
                                         "start_dt":start_dt,"end_dt":end_dt,"runtime":runtime,
                                         "screen":"","source":"seoulart",
                                         "show_type":"","program":""})
                        except:
                            continue
                next_tr=next_tr.find_next_sibling()
    return rows

def fetch_kofa(cinema):
    rows=[]
    url="https://www.koreafilm.or.kr/cinematheque/schedule"
    try:
        res=requests.get(url, timeout=10)
        soup=BeautifulSoup(res.text,"html.parser")
    except:
        return rows
    today=datetime.today()
    month=today.month
    prev_day=None
    year=today.year
    blocks=soup.find_all("dl","list-day-1")
    for block in blocks:
        date_tag=block.find("dt","txt-day")
        if not date_tag: continue
        txt=date_tag.text.strip()
        if "." not in txt: continue
        day=int(txt.split(".")[0])
        if prev_day and day<prev_day: month+=1
        prev_day=day
        date_obj=datetime(year,month,day)
        for s in block.select("ul.list-detail-1"):
            start_tag=s.select_one(".txt-time")
            start=start_tag.get_text(strip=True) if start_tag else None
            runtime_tag=s.select_one(".min")
            if runtime_tag:
                for strong in runtime_tag.find_all("strong"):
                    strong.decompose()
                runtime_text=runtime_tag.get_text(strip=True).replace("분","")
                runtime=int(runtime_text) if runtime_text.isdigit() else None
            else:
                runtime=None
            end_dt=compute_end_dt(make_datetime(date_obj,start),runtime)
            title_tag=s.select_one(".txt-1 a")
            title=title_tag.get_text(strip=True) if title_tag else ""
            screen_tag=s.select_one(".txt-room")
            screen=screen_tag.get_text(strip=True) if screen_tag else ""
            type_tag=s.select_one(".fomat")
            show_type=type_tag.get_text(strip=True)[4:] if type_tag else ""
            program_tag=s.select_one(".layer-txt-1")
            program=program_tag.get_text(strip=True) if program_tag else ""
            rows.append({
                "cinema":cinema["name"], "movie":title,
                "start_dt":make_datetime(date_obj,start), "end_dt":end_dt,
                "runtime":runtime, "screen":screen, "source":"kofa",
                "show_type":show_type, "program":program
            })
    return rows

# ── 크롤링 ───────────────────────────────────────────
def run_crawl():
    log.info("크롤링 시작")
    start_time = time.time()
    all_rows = []
    today = datetime.today()
    for cinema in CINEMAS:
        src = cinema["source"]
        try:
            if src=="dtryx":      rows=fetch_dtryx(cinema,today,days=14)
            elif src=="moviee":   rows=fetch_moviee(cinema,today,days=14)
            elif src=="seoulart": rows=fetch_seoulart(cinema)
            elif src=="kofa":     rows=fetch_kofa(cinema)
            else: rows=[]
            all_rows += rows
            log.info(f"  {cinema['name']} 수집 완료: {len(rows)}건")
        except Exception as e:
            log.error(f"  {cinema['name']} 오류: {e}")
    all_rows = sorted(all_rows, key=lambda x: (x['start_dt'] or datetime.max, x['end_dt'] or datetime.max))
    save_to_db(all_rows)
    log.info(f"총 {len(all_rows)}건 저장 완료, 처리시간: {time.time()-start_time:.2f}초")
    return len(all_rows)



# ── API ──────────────────────────────────────────────
@app.route("/")
def index():
    with open(os.path.join(app.template_folder, "index.html"), encoding="utf-8") as f:
        return f.read()

@app.route("/api/screenings")
def api_screenings():
    df = request.args.get("date_from", date.today().isoformat())
    dt = request.args.get("date_to",   date.today().isoformat())
    tf = request.args.get("time_from", "00:00")
    tt = request.args.get("time_to",   "23:59")
    cinemas = request.args.getlist("cinema")
    mq = request.args.get("movie", "").strip()
    conn = get_db(); cur = conn.cursor()
    sql = """
        SELECT s.*, COALESCE(m.poster_url, '') AS poster_url
        FROM screenings s
        LEFT JOIN movies m ON m.title = s.movie
        WHERE s.start_dt::date BETWEEN %s AND %s
          AND s.start_dt::time >= %s::time
          AND s.start_dt::time <= %s::time
    """
    params = [df, dt, tf, tt]
    if cinemas:
        sql += f" AND s.cinema = ANY(%s)"; params.append(cinemas)
    if mq:
        sql += " AND s.movie ILIKE %s"; params.append(f"%{mq}%")
    sql += " ORDER BY s.start_dt ASC"
    cur.execute(sql, params)
    rows = [dict(r) for r in cur.fetchall()]
    cur.close(); conn.close()
    return jsonify(rows)

@app.route("/api/movies")
def api_movies():
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT title, director, synopsis, poster_url FROM movies WHERE poster_url != '' ORDER BY title")
    rows = [dict(r) for r in cur.fetchall()]
    cur.close(); conn.close()
    return jsonify(rows)

@app.route("/api/recommended", methods=["GET"])
def api_recommended_get():
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT id, title, description, awards, created_at FROM recommended ORDER BY created_at DESC")
    rows = [dict(r) for r in cur.fetchall()]
    cur.close(); conn.close()
    return jsonify(rows)

@app.route("/api/recommended", methods=["POST"])
def api_recommended_post():
    if not session.get("admin"):
        return jsonify({"error": "unauthorized"}), 401
    data = request.get_json()
    title = (data.get("title") or "").strip()
    description = (data.get("description") or "").strip()
    awards = (data.get("awards") or "").strip()
    if not title:
        return jsonify({"error": "title required"}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute("""
        INSERT INTO recommended (title, description, awards) VALUES (%s, %s, %s)
        ON CONFLICT (title) DO UPDATE SET description=EXCLUDED.description, awards=EXCLUDED.awards
    """, (title, description, awards))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True})

@app.route("/api/recommended/<int:rid>", methods=["DELETE"])
def api_recommended_delete(rid):
    if not session.get("admin"):
        return jsonify({"error": "unauthorized"}), 401
    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM recommended WHERE id=%s", (rid,))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True})

@app.route("/api/cinemas")
def api_cinemas():
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT DISTINCT cinema FROM screenings")
    names = [r["cinema"] for r in cur.fetchall()]
    cur.close(); conn.close()
    ordered = [c for c in CINEMA_ORDER if c in names] + [c for c in names if c not in CINEMA_ORDER]
    return jsonify(ordered)

@app.route("/api/stats")
def api_stats():
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT COUNT(*) AS c FROM screenings"); total = cur.fetchone()["c"]
    cur.execute("SELECT COUNT(DISTINCT cinema) AS c FROM screenings"); cc = cur.fetchone()["c"]
    cur.execute("SELECT COUNT(DISTINCT movie) AS c FROM screenings"); mc = cur.fetchone()["c"]
    cur.execute("SELECT value FROM meta WHERE key='last_updated'"); lu = cur.fetchone()
    cur.execute("SELECT MIN(start_dt::date) AS mn, MAX(start_dt::date) AS mx FROM screenings WHERE start_dt IS NOT NULL")
    dr = cur.fetchone()
    cur.close(); conn.close()
    return jsonify({"total": total, "cinemas": cc, "movies": mc,
                    "last_updated": lu["value"] if lu else None,
                    "date_min": str(dr["mn"]) if dr and dr["mn"] else None,
                    "date_max": str(dr["mx"]) if dr and dr["mx"] else None})

@app.route("/api/cinema/<name>")
def api_cinema_detail(name):
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT * FROM cinemas WHERE name = %s", (name,))
    row = cur.fetchone()
    cinema = dict(row) if row else {"name": name}
    cur.execute(
        "SELECT * FROM screenings WHERE cinema = %s AND start_dt::date >= CURRENT_DATE ORDER BY start_dt ASC",
        (name,)
    )
    screenings = [dict(r) for r in cur.fetchall()]
    cur.execute(
        "SELECT MIN(start_dt::date) AS mn, MAX(start_dt::date) AS mx FROM screenings WHERE cinema = %s AND start_dt IS NOT NULL",
        (name,)
    )
    dr = cur.fetchone()
    cur.close(); conn.close()
    return jsonify({
        "cinema": cinema,
        "screenings": screenings,
        "date_max": str(dr["mx"]) if dr and dr["mx"] else None,
    })

@app.route("/api/cinemas/all")
def api_cinemas_all():
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT * FROM cinemas ORDER BY name")
    rows = [dict(r) for r in cur.fetchall()]
    cur.close(); conn.close()
    return jsonify(rows)

@app.route("/cinema/<name>")
def cinema_page(name):
    with open(os.path.join(app.template_folder, "cinema.html"), encoding="utf-8") as f:
        return f.read()

@app.route("/api/refresh", methods=["POST"])
def api_refresh():
    secret = request.headers.get("X-Secret", "")
    if os.environ.get("REFRESH_SECRET") and secret != os.environ["REFRESH_SECRET"]:
        return jsonify({"error": "unauthorized"}), 401
    threading.Thread(target=run_crawl, daemon=True).start()
    return jsonify({"status": "started"})

@app.route("/api/export/excel")
def export_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    df = request.args.get("date_from", date.today().isoformat())
    dt = request.args.get("date_to",   date.today().isoformat())
    cinemas = request.args.getlist("cinema"); mq = request.args.get("movie", "").strip()
    conn = get_db(); cur = conn.cursor()
    sql = "SELECT * FROM screenings WHERE start_dt::date BETWEEN %s AND %s"
    params = [df, dt]
    if cinemas:
        sql += " AND cinema = ANY(%s)"; params.append(cinemas)
    if mq:
        sql += " AND movie ILIKE %s"; params.append(f"%{mq}%")
    cur.execute(sql + " ORDER BY start_dt", params)
    rows = cur.fetchall(); cur.close(); conn.close()

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "시간표"
    headers = ["극장","영화제목","날짜","시작","종료","런타임(분)","상영관","상영유형","프로그램"]
    thin = Side(style="thin", color="DDDDDD"); bdr = Border(left=thin,right=thin,top=thin,bottom=thin)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = PatternFill("solid", fgColor="1A1A2E")
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.alignment = Alignment(horizontal="center"); c.border = bdr
    ws.row_dimensions[1].height = 24
    palette = ["EFF6FF","FFF7ED","F0FDF4","FFF1F2","F5F3FF","ECFEFF","FFFBEB","F0F9FF"]
    colors = {}; pidx = 0
    for ri, r in enumerate(rows, 2):
        if r["cinema"] not in colors:
            colors[r["cinema"]] = palette[pidx % len(palette)]; pidx += 1
        fill = PatternFill("solid", fgColor=colors[r["cinema"]])
        sdt = r["start_dt"] or ""
        for col, v in enumerate([r["cinema"], r["movie"], str(sdt)[:10], str(sdt)[11:16],
                str(r["end_dt"] or "")[11:16], r["runtime"], r["screen"],
                r["show_type"], r["program"]], 1):
            c = ws.cell(row=ri, column=col, value=v)
            c.fill = fill; c.border = bdr; c.alignment = Alignment(vertical="center")
    for col, w in zip("ABCDEFGHI", [16,36,12,10,10,10,14,16,24]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"시간표_{df}_{dt}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ── 시작 ─────────────────────────────────────────────
ensure_db()

def initial_crawl_if_empty():
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("SELECT COUNT(*) AS c FROM screenings")
        count = cur.fetchone()["c"]; cur.close(); conn.close()
        if count == 0:
            log.info("DB 비어있음 → 최초 수집 시작 (백그라운드)")
            threading.Thread(target=run_crawl, daemon=True).start()
        else:
            log.info(f"기존 DB 사용: {count}건")
    except Exception as e:
        log.error(f"DB 초기화 오류: {e}")

initial_crawl_if_empty()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)

# ── 관리자 ────────────────────────────────────────────
@app.route("/admin", methods=["GET", "POST"])
def admin():
    if not ADMIN_PASSWORD:
        return "관리자 페이지가 비활성화되어 있습니다. ADMIN_PASSWORD 환경변수를 설정해주세요.", 503
    ip = _get_ip()
    if request.method == "POST" and "password" in request.form:
        if _is_locked(ip):
            rec = _login_attempts.get(ip, {})
            remain = int((rec["locked_until"] - datetime.now()).total_seconds() // 60) + 1 if rec.get("locked_until") else LOCKOUT_MINUTES
            return render_template_string(ADMIN_LOGIN_TEMPLATE,
                error=f"로그인 시도가 너무 많습니다. {remain}분 후 다시 시도해주세요.")
        if request.form["password"] == ADMIN_PASSWORD:
            _clear_attempts(ip)
            session["admin"] = True
            return redirect(url_for("admin_dashboard"))
        _record_failure(ip)
        rec = _login_attempts.get(ip, {})
        if rec.get("locked_until"):
            error = f"비밀번호가 틀렸습니다. 시도 횟수 초과로 {LOCKOUT_MINUTES}분간 잠금됩니다."
        else:
            left = MAX_ATTEMPTS - rec.get("count", 0)
            error = f"비밀번호가 틀렸습니다. (남은 시도: {left}회)"
        return render_template_string(ADMIN_LOGIN_TEMPLATE, error=error)
    if session.get("admin"):
        return redirect(url_for("admin_dashboard"))
    return render_template_string(ADMIN_LOGIN_TEMPLATE, error=None)

@app.route("/admin/dashboard")
def admin_dashboard():
    if not session.get("admin"):
        return redirect(url_for("admin"))
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT * FROM cinemas ORDER BY name")
    cinemas = [dict(r) for r in cur.fetchall()]
    cur.execute("SELECT * FROM movies ORDER BY title")
    movies = [dict(r) for r in cur.fetchall()]
    cur.close(); conn.close()
    return render_template_string(ADMIN_DASHBOARD_TEMPLATE, cinemas=cinemas, movies=movies)

@app.route("/admin/logout")
def admin_logout():
    session.pop("admin", None)
    return redirect(url_for("admin"))

@app.route("/admin/cinema/save", methods=["POST"])
def admin_cinema_save():
    if not session.get("admin"): return redirect(url_for("admin"))
    d = request.form
    conn = get_db(); cur = conn.cursor()
    cur.execute("""
        INSERT INTO cinemas (name,address,url,phone,description,is_free,note,updated_at)
        VALUES (%s,%s,%s,%s,%s,%s,%s,NOW())
        ON CONFLICT (name) DO UPDATE SET
            address=EXCLUDED.address, url=EXCLUDED.url, phone=EXCLUDED.phone,
            description=EXCLUDED.description, is_free=EXCLUDED.is_free,
            note=EXCLUDED.note, updated_at=NOW()
    """, (d["name"], d.get("address",""), d.get("url",""), d.get("phone",""),
          d.get("description",""), d.get("is_free","") == "on", d.get("note","")))
    conn.commit(); cur.close(); conn.close()
    return redirect(url_for("admin_dashboard") + "#cinemas")

@app.route("/admin/movie/save", methods=["POST"])
def admin_movie_save():
    if not session.get("admin"): return redirect(url_for("admin"))
    d = request.form
    conn = get_db(); cur = conn.cursor()
    cur.execute("""
        INSERT INTO movies (title,title_en,director,country,year,runtime,synopsis,poster_url,kobis_url,kofa_url,note,updated_at)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())
        ON CONFLICT (title) DO UPDATE SET
            title_en=EXCLUDED.title_en, director=EXCLUDED.director, country=EXCLUDED.country,
            year=EXCLUDED.year, runtime=EXCLUDED.runtime, synopsis=EXCLUDED.synopsis,
            poster_url=EXCLUDED.poster_url, kobis_url=EXCLUDED.kobis_url,
            kofa_url=EXCLUDED.kofa_url, note=EXCLUDED.note, updated_at=NOW()
    """, (d["title"], d.get("title_en",""), d.get("director",""), d.get("country",""),
          int(d["year"]) if d.get("year") else None,
          int(d["runtime"]) if d.get("runtime") else None,
          d.get("synopsis",""), d.get("poster_url",""), d.get("kobis_url",""),
          d.get("kofa_url",""), d.get("note","")))
    conn.commit(); cur.close(); conn.close()
    return redirect(url_for("admin_dashboard") + "#movies")

@app.route("/admin/movie/delete/<int:mid>", methods=["POST"])
def admin_movie_delete(mid):
    if not session.get("admin"): return redirect(url_for("admin"))
    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM movies WHERE id = %s", (mid,))
    conn.commit(); cur.close(); conn.close()
    return redirect(url_for("admin_dashboard") + "#movies")

# ── 시작 ─────────────────────────────────────────────
ensure_db()

def initial_crawl_if_empty():
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("SELECT COUNT(*) AS c FROM screenings")
        count = cur.fetchone()["c"]; cur.close(); conn.close()
        if count == 0:
            log.info("DB 비어있음 → 최초 수집 시작 (백그라운드)")
            threading.Thread(target=run_crawl, daemon=True).start()
        else:
            log.info(f"기존 DB 사용: {count}건")
    except Exception as e:
        log.error(f"DB 초기화 오류: {e}")

initial_crawl_if_empty()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)

# ── 관리자 HTML 템플릿 ─────────────────────────────────
ADMIN_LOGIN_TEMPLATE = """<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8"><title>관리자 로그인</title>
<link href="https://fonts.googleapis.com/css2?family=Pretendard:wght@400;600;700&display=swap" rel="stylesheet">
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Pretendard',sans-serif;background:#F7F8FA;display:flex;align-items:center;justify-content:center;min-height:100vh}
.card{background:#fff;border:1px solid #E2E5EB;border-radius:12px;padding:40px;width:360px;box-shadow:0 4px 16px rgba(0,0,0,.08)}
.logo{text-align:center;font-size:12px;color:#9CA3AF;margin-bottom:24px;letter-spacing:.04em}
h2{font-size:20px;font-weight:700;margin-bottom:24px;text-align:center}
label{font-size:13px;font-weight:600;color:#374151;display:block;margin-bottom:6px}
input{width:100%;padding:10px 14px;border:1.5px solid #D1D5DB;border-radius:7px;font-size:14px;font-family:inherit;outline:none}
input:focus{border-color:#336600}
.btn{width:100%;margin-top:16px;padding:11px;background:#336600;color:#fff;border:none;border-radius:7px;font-size:14px;font-weight:700;cursor:pointer;font-family:inherit}
.btn:hover{background:#2a5200}
.error{color:#DC2626;font-size:13px;margin-top:10px;text-align:center}
</style>
</head>
<body>
<div class="card">
  <div class="logo">SEOUL // INDIE CINEMA</div>
  <h2>관리자 로그인</h2>
  <form method="POST">
    <label>비밀번호</label>
    <input type="password" name="password" autofocus placeholder="비밀번호 입력">
    <button class="btn" type="submit">로그인</button>
    {% if error %}<div class="error">{{ error }}</div>{% endif %}
  </form>
</div>
</body>
</html>"""

ADMIN_DASHBOARD_TEMPLATE = """<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>관리자 — SEOUL INDIE CINEMA</title>
<link href="https://fonts.googleapis.com/css2?family=Pretendard:wght@300;400;500;600;700&display=swap" rel="stylesheet">
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Pretendard',sans-serif;background:#F7F8FA;color:#111827}
header{background:#fff;border-bottom:1px solid #E2E5EB;padding:0 24px;position:sticky;top:0;z-index:100}
.header-inner{max-width:1200px;margin:0 auto;height:56px;display:flex;align-items:center;justify-content:space-between}
.logo{font-size:13px;font-weight:500;color:#336600;letter-spacing:.04em}
.hbtns{display:flex;gap:8px}
.btn{padding:7px 16px;border-radius:7px;border:none;cursor:pointer;font-size:13px;font-weight:600;font-family:inherit;transition:all .15s;text-decoration:none;display:inline-flex;align-items:center}
.btn-primary{background:#336600;color:#fff}.btn-primary:hover{background:#2a5200}
.btn-secondary{background:#fff;color:#374151;border:1.5px solid #D1D5DB}.btn-secondary:hover{border-color:#111827;color:#111827}
.btn-danger{background:#fff;color:#DC2626;border:1.5px solid #FECACA;padding:4px 10px;font-size:11px}.btn-danger:hover{background:#FEF2F2}
.main{max-width:1200px;margin:0 auto;padding:28px 24px 60px}
.tabs{display:flex;border-bottom:2px solid #E2E5EB;margin-bottom:24px}
.tab{padding:10px 20px;font-size:14px;font-weight:600;cursor:pointer;color:#6B7280;border-bottom:2px solid transparent;margin-bottom:-2px;transition:all .12s}
.tab.active{color:#111827;border-bottom-color:#111827}
.section{display:none}.section.active{display:block}
.card{background:#fff;border:1px solid #E2E5EB;border-radius:12px;padding:24px 28px;margin-bottom:20px;box-shadow:0 1px 3px rgba(0,0,0,.06)}
.card-title{font-size:15px;font-weight:700;margin-bottom:20px;padding-bottom:10px;border-bottom:1.5px solid #F0F2F5}
.form-grid{display:grid;grid-template-columns:1fr 1fr;gap:14px}
.fg{display:flex;flex-direction:column;gap:5px}
.fg.full{grid-column:1/-1}
label{font-size:11px;font-weight:700;color:#6B7280;text-transform:uppercase;letter-spacing:.04em}
input,select,textarea{padding:9px 12px;border:1.5px solid #D1D5DB;border-radius:7px;font-size:14px;font-family:inherit;outline:none;background:#fff;width:100%}
input:focus,select:focus,textarea:focus{border-color:#336600;box-shadow:0 0 0 3px rgba(51,102,0,.08)}
textarea{resize:vertical;min-height:80px}
.cb-row{display:flex;align-items:center;gap:8px;margin-top:4px}
.cb-row input{width:auto;padding:0}
.cb-row label{text-transform:none;font-size:14px;font-weight:500;color:#111827;letter-spacing:0}
table{width:100%;border-collapse:separate;border-spacing:0;font-size:13px}
thead{background:#F8FAFC}
th{padding:10px 14px;text-align:left;font-weight:700;font-size:11px;color:#6B7280;border-bottom:1.5px solid #E2E5EB;white-space:nowrap}
td{padding:10px 14px;border-bottom:1px solid #F0F2F5;vertical-align:middle}
tr:last-child td{border-bottom:none}
tr:hover td{background:#FAFAFA}
.badge{display:inline-block;font-size:11px;font-weight:700;padding:2px 8px;border-radius:4px}
.badge-free{background:#D1FAE5;color:#065F46}
.badge-paid{background:#F1F5F9;color:#374151}
a{color:#336600}
.save-btn{margin-top:16px}
</style>
</head>
<body>
<header>
  <div class="header-inner">
    <a href="/" class="logo">SEOUL // INDIE CINEMA</a>
    <div class="hbtns">
      <a href="/" class="btn btn-secondary">← 사이트</a>
      <a href="/admin/logout" class="btn btn-secondary">로그아웃</a>
    </div>
  </div>
</header>

<div class="main">
  <div class="tabs">
    <div class="tab active" onclick="switchTab('cinemas')">🎬 극장 관리</div>
    <div class="tab" onclick="switchTab('movies')">🎞 영화 관리</div>
    <div class="tab" onclick="switchTab('recommended')">★ 추천 관리</div>
  </div>

  <div id="tab-cinemas" class="section active">
    <div class="card" id="cinemas">
      <div class="card-title">극장 정보 입력 / 수정</div>
      <form method="POST" action="/admin/cinema/save">
        <div class="form-grid">
          <div class="fg"><label>극장명 *</label><input type="text" name="name" required placeholder="예: KU시네마테크"></div>
          <div class="fg"><label>홈페이지 URL</label><input type="url" name="url" placeholder="https://..."></div>
          <div class="fg"><label>주소</label><input type="text" name="address" placeholder="서울시 ..."></div>
          <div class="fg"><label>전화번호</label><input type="text" name="phone" placeholder="02-..."></div>
          <div class="fg full"><label>소개</label><textarea name="description" placeholder="극장 소개 및 특징..."></textarea></div>
          <div class="fg full"><label>메모 (내부용)</label><input type="text" name="note"></div>
          <div class="fg full"><div class="cb-row"><input type="checkbox" name="is_free" id="is_free"><label for="is_free">무료 상영관</label></div></div>
        </div>
        <div class="save-btn"><button type="submit" class="btn btn-primary">저장</button></div>
      </form>
    </div>
    <div class="card">
      <div class="card-title">등록된 극장 목록 ({{ cinemas|length }}개)</div>
      {% if cinemas %}
      <div style="overflow-x:auto"><table>
        <thead><tr><th>극장명</th><th>주소</th><th>전화</th><th>홈페이지</th><th>무료</th><th>소개</th></tr></thead>
        <tbody>
        {% for c in cinemas %}<tr>
          <td><strong>{{ c.name }}</strong></td>
          <td style="color:#6B7280">{{ c.address or "—" }}</td>
          <td style="color:#6B7280">{{ c.phone or "—" }}</td>
          <td>{% if c.url %}<a href="{{ c.url }}" target="_blank">링크</a>{% else %}—{% endif %}</td>
          <td>{% if c.is_free %}<span class="badge badge-free">무료</span>{% else %}<span class="badge badge-paid">유료</span>{% endif %}</td>
          <td style="font-size:12px;color:#6B7280;max-width:200px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis">{{ c.description or "—" }}</td>
        </tr>{% endfor %}
        </tbody>
      </table></div>
      {% else %}<p style="color:#9CA3AF;font-size:14px">등록된 극장이 없습니다.</p>{% endif %}
    </div>
  </div>

  <div id="tab-movies" class="section">
    <div class="card" id="movies">
      <div class="card-title">영화 정보 입력 / 수정</div>
      <form method="POST" action="/admin/movie/save">
        <div class="form-grid">
          <div class="fg"><label>영화 제목 (한글) *</label><input type="text" name="title" required placeholder="예: 기억의 땅"></div>
          <div class="fg"><label>영어 제목</label><input type="text" name="title_en" placeholder="예: Land of Memory"></div>
          <div class="fg"><label>감독</label><input type="text" name="director"></div>
          <div class="fg"><label>국가</label><input type="text" name="country" placeholder="예: 한국, 프랑스"></div>
          <div class="fg"><label>제작연도</label><input type="number" name="year" placeholder="2024" min="1900" max="2099"></div>
          <div class="fg"><label>러닝타임 (분)</label><input type="number" name="runtime" placeholder="90"></div>
          <div class="fg full"><label>시놉시스</label><textarea name="synopsis" style="min-height:100px"></textarea></div>
          <div class="fg"><label>포스터 URL</label><input type="url" name="poster_url" placeholder="https://..."></div>
          <div class="fg"><label>KOBIS URL</label><input type="url" name="kobis_url" placeholder="https://kobis.or.kr/..."></div>
          <div class="fg"><label>KOFA URL</label><input type="url" name="kofa_url" placeholder="https://www.koreafilm.or.kr/..."></div>
          <div class="fg full"><label>메모</label><input type="text" name="note"></div>
        </div>
        <div class="save-btn"><button type="submit" class="btn btn-primary">저장</button></div>
      </form>
    </div>
    <div class="card">
      <div class="card-title">등록된 영화 목록 ({{ movies|length }}편)</div>
      {% if movies %}
      <div style="overflow-x:auto"><table>
        <thead><tr><th>제목</th><th>감독</th><th>국가/연도</th><th>런타임</th><th>KOFA</th><th>삭제</th></tr></thead>
        <tbody>
        {% for m in movies %}<tr>
          <td><strong>{{ m.title }}</strong>{% if m.title_en %}<br><span style="font-size:12px;color:#6B7280">{{ m.title_en }}</span>{% endif %}</td>
          <td>{{ m.director or "—" }}</td>
          <td style="white-space:nowrap">{{ m.country or "" }}{% if m.year %} {{ m.year }}{% endif %}</td>
          <td>{{ (m.runtime|string ~ "분") if m.runtime else "—" }}</td>
          <td>{% if m.kofa_url %}<a href="{{ m.kofa_url }}" target="_blank">링크</a>{% else %}—{% endif %}</td>
          <td><form method="POST" action="/admin/movie/delete/{{ m.id }}" onsubmit="return confirm('삭제할까요?')"><button type="submit" class="btn btn-danger">삭제</button></form></td>
        </tr>{% endfor %}
        </tbody>
      </table></div>
      {% else %}<p style="color:#9CA3AF;font-size:14px">등록된 영화가 없습니다.</p>{% endif %}
    </div>
  </div>

  <div id="tab-recommended" class="section">
    <div class="card">
      <div class="card-title">★ 관리자 추천 영화 등록 / 수정</div>
      <p style="font-size:13px;color:#6B7280;margin-bottom:16px;">
        영화 제목은 상영 DB에 등록된 제목과 <strong>정확히 일치</strong>해야 합니다.<br>
        수상 내역은 쉼표(,)로 구분해 입력하면 자동으로 칩 형태로 표시됩니다.
      </p>
      <div class="form-grid">
        <div class="fg"><label>영화 제목 *</label>
          <input type="text" id="rec-title" placeholder="예: 사랑의 기억"></div>
        <div class="fg"><label>수상 내역 (쉼표 구분)</label>
          <input type="text" id="rec-awards" placeholder="예: 2025 칸:황금종려상 수상, 2024 아카데미:각본상 후보"></div>
        <div class="fg full"><label>소개글</label>
          <textarea id="rec-desc" style="min-height:80px" placeholder="관람 포인트나 추천 이유를 자유롭게 적어주세요"></textarea></div>
      </div>
      <div class="save-btn"><button class="btn btn-primary" onclick="addRecommended()">저장</button></div>
    </div>
    <div class="card">
      <div class="card-title">현재 추천 목록</div>
      <div id="rec-list"><p style="color:#9CA3AF;font-size:14px">로딩 중…</p></div>
    </div>
  </div>
</div>

<script>
const TABS = ["cinemas","movies","recommended"];
function switchTab(name) {
  document.querySelectorAll(".tab").forEach((t,i)=>t.classList.toggle("active",TABS[i]===name));
  document.querySelectorAll(".section").forEach(s=>s.classList.remove("active"));
  document.getElementById("tab-"+name).classList.add("active");
  if(name==="recommended") loadRecommended();
}
if(location.hash==="#movies") switchTab("movies");
if(location.hash==="#recommended") switchTab("recommended");

async function loadRecommended(){
  const res = await fetch("/api/recommended");
  const rows = await res.json();
  const el = document.getElementById("rec-list");
  if(!rows.length){ el.innerHTML='<p style="color:#9CA3AF;font-size:14px">등록된 추천이 없습니다.</p>'; return; }
  el.innerHTML = `<table><thead><tr>
    <th>영화 제목</th><th>수상 내역</th><th>소개글</th><th>등록일</th><th>수정</th><th>삭제</th>
  </tr></thead><tbody>
    ${rows.map(r=>`<tr>
      <td><strong>${r.title}</strong></td>
      <td style="font-size:12px;color:#6B7280">${(r.awards||"").split(",").filter(Boolean).map(a=>{
        const [fest,...rest]=a.trim().split(":");
        return `<span style="display:inline-flex;align-items:center;margin:1px 4px 1px 0;">
          <span style="background:#555;color:#fff;font-size:10px;font-weight:700;padding:1px 6px;border-radius:3px 0 0 3px;">${fest.trim()}</span
          ><span style="font-size:10px;color:#111;padding:1px 5px;border:1px solid #ddd;border-left:none;border-radius:0 3px 3px 0;">${(rest.join("|")||"").trim()}</span>
        </span>`;}).join("") || "—"}</td>
      <td style="font-size:12px;color:#6B7280;max-width:200px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">${r.description||"—"}</td>
      <td style="font-size:12px;color:#9CA3AF;white-space:nowrap">${(r.created_at||"").slice(0,10)}</td>
      <td><button class="btn btn-secondary" style="font-size:11px;padding:3px 10px;"
        data-title="${r.title}" data-awards="${r.awards||''}" data-desc="${(r.description||'').replace(/"/g,'&quot;')}"
        onclick="editRecommended(this)">수정</button></td>
      <td><button class="btn btn-danger" onclick="deleteRecommended(${r.id})">삭제</button></td>
    </tr>`).join("")}
  </tbody></table>`;
}
function editRecommended(btn){
  document.getElementById("rec-title").value = btn.dataset.title;
  document.getElementById("rec-awards").value = btn.dataset.awards;
  document.getElementById("rec-desc").value = btn.dataset.desc;
  document.getElementById("rec-title").scrollIntoView({behavior:"smooth", block:"center"});
}
async function addRecommended(){
  const title = document.getElementById("rec-title").value.trim();
  const awards = document.getElementById("rec-awards").value.trim();
  const description = document.getElementById("rec-desc").value.trim();
  if(!title){ alert("영화 제목을 입력하세요."); return; }
  const res = await fetch("/api/recommended",{
    method:"POST", headers:{"Content-Type":"application/json"},
    body: JSON.stringify({title, description, awards})
  });
  if(res.ok){
    document.getElementById("rec-title").value="";
    document.getElementById("rec-awards").value="";
    document.getElementById("rec-desc").value="";
    loadRecommended();
  } else { const d=await res.json(); alert(d.error||"오류 발생"); }
}
async function deleteRecommended(id){
  if(!confirm("삭제할까요?")) return;
  await fetch(`/api/recommended/${id}`,{method:"DELETE"});
  loadRecommended();
}
</script>
</body>
</html>"""
